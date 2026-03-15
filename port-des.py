import os
import sys
import logging
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from netmiko import ConnectHandler
from dotenv import load_dotenv

# ============================================================
# 1. SETUP & DEBUG LOGGING
# ============================================================
load_dotenv()

# This logs the "Netmiko internal talk" to a file
logging.basicConfig(filename='session_debug.txt', level=logging.DEBUG)
logger = logging.getLogger("netmiko")

USERNAME = os.getenv("username")
PASSWORD = os.getenv("passwordAD")
MAX_SWITCH_CONCURRENCY = 3 # Lowered slightly for cleaner debug output

def process_switch_queue(switch_ip, tasks):
    device = {
        'device_type': 'aruba_osswitch',
        'host': switch_ip,
        'username': USERNAME,
        'password': PASSWORD,
        'conn_timeout': 30,
        'global_delay_factor': 2.5, # Slowed down for stability
        'session_log': f'log_{switch_ip}.txt' # Creates a specific log for each switch
    }
    
    results = []
    try:
        with ConnectHandler(**device) as net_connect:
            print(f"[*] {switch_ip}: Connection established. Entering support-mode...")
            
            # Step 1: Support Mode
            out = net_connect.send_command("aruba-central support-mode", expect_string=r"y/n|#")
            if "y/n" in out:
                net_connect.send_command("y", expect_string=r"#")
            
            # Step 2: Config Mode
            net_connect.send_command("conf t", expect_string=r"\(config\)#")
            print(f"[*] {switch_ip}: In Config Mode.")

            for row_idx, port, description in tasks:
                try:
                    print(f"    --> {switch_ip} [Row {row_idx}]: Setting port {port}")
                    
                    # Step 3: Interface
                    net_connect.send_command(f"interface {port}", expect_string=r"\(config-if\)#")
                    
                    # Step 4: Description
                    # We use send_command_timing here to handle slower switch responses
                    net_connect.send_command(f"description \"{description}\"")
                    
                    # Step 5: Exit back to global config
                    net_connect.send_command("exit", expect_string=r"\(config\)#")
                    
                    results.append(f"[ROW {row_idx}] SUCCESS: {switch_ip} port {port} set to {description}")
                
                except Exception as row_error:
                    print(f"    [!] Error on Row {row_idx}: {row_error}")
                    results.append(f"[ROW {row_idx}] FAILED: {switch_ip} error: Row Timeout")
                    # Try to reset prompt state if one row fails
                    net_connect.send_command("\x03") # Send Ctrl+C to reset
                    net_connect.send_command("conf t")
            
            print(f"[*] {switch_ip}: Saving memory...")
            net_connect.send_command("write memory", expect_string=r"#")
                
    except Exception as e:
        print(f"    [CRITICAL] {switch_ip} Connection Lost: {e}")
        for row_idx, port, description in tasks:
            if not any(f"[ROW {row_idx}]" in r for r in results):
                results.append(f"[ROW {row_idx}] FAILED: {switch_ip} connection error")
    
    return results

# ============================================================
# 2. MAIN EXECUTION (No changes to logic here)
# ============================================================
def main():
    if len(sys.argv) < 2:
        print("\n[!] Usage: python3 update-descriptions.py <filename.xlsx>")
        return

    file_path = sys.argv[1]
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    switch_batches = {}
    for row_idx in range(4, ws.max_row + 1):
        bldg = ws.cell(row=row_idx, column=1).value
        cr = ws.cell(row=row_idx, column=2).value
        num = ws.cell(row=row_idx, column=3).value
        sw_ip = ws.cell(row=row_idx, column=8).value
        port = ws.cell(row=row_idx, column=9).value

        if sw_ip and port and bldg and cr and num:
            ip = str(sw_ip).strip()
            if ip not in switch_batches: switch_batches[ip] = []
            switch_batches[ip].append((row_idx, str(port).strip(), f"{bldg}/{cr}/{num}"))

    all_final_results = []
    with ThreadPoolExecutor(max_workers=MAX_SWITCH_CONCURRENCY) as executor:
        future_to_switch = {executor.submit(process_switch_queue, ip, tasks): ip for ip, tasks in switch_batches.items()}
        for future in future_to_switch:
            all_final_results.extend(future.result())

    print("\n" + "="*75)
    print("FINISHED. CHECK log_X.X.X.X.txt FOR DETAILED SWITCH REPLIES.")
    print("="*75)

if __name__ == "__main__":
    main()