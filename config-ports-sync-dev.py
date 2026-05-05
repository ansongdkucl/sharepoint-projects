import errno
import os
import re
import sys
import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from netmiko import ConnectHandler
from openpyxl import load_workbook

load_dotenv()

USERNAME = os.getenv("username")
PASSWORD = os.getenv("passwordAD")
TEAMS_WEBHOOK_URL = os.getenv("TEAMS_WEBHOOK_URL", "")

ROUTER_IP = "aruba-dist"
ROUTER_DEVICE_TYPE = "aruba_aoscx"

RUN_ACTOR = (
    os.getenv("GITHUB_ACTOR")
    or os.getenv("USER")
    or os.getenv("USERNAME")
    or "unknown"
)

RUN_SOURCE = (
    "GitHub Actions"
    if os.getenv("GITHUB_ACTIONS", "").lower() == "true"
    else "Manual"
)

# ============================================================
# FILE LIST - comment out any file you do not want processed
# Adjust names here to match the exact filenames on disk
# ============================================================
ONEDRIVE_DIR = Path(
    "/mnt/c/Users/cceadan/OneDrive - University College London/Estates IT - Project Documentation - Patching Schedule"
)

DOWNLOADS_DIR = Path(
    "/mnt/c/Users/cceadan/Downloads"
)

FILES_TO_CHECK = [
    #"90TCR - Daniel Test.xlsx"
   "90 TCR - Level 3A Patching Schedule.xlsx",

   # "90TCR - Level 2B - Patching Schedule.xlsx",
   # "90TCR - Level 3B - Patching Schedule.xlsx",
    # "another file.xlsx",
]

SEARCH_DIRS = [ONEDRIVE_DIR, DOWNLOADS_DIR]


# ============================================================
# Helpers
# ============================================================
def now_str():
    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def ordinal(day):
    if 10 <= day % 100 <= 20:
        return f"{day}th"
    return f"{day}{['th','st','nd','rd','th','th','th','th','th','th'][day % 10]}"


def now_friendly():
    dt = datetime.now().astimezone()
    return dt.strftime(f"%H:%M %A {ordinal(dt.day)} %B")


def log(msg):
    print(msg, flush=True)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value):
    text = str(value or "").replace("\n", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def normalize_mac(value):
    return re.sub(r"[^0-9a-fA-F]", "", str(value or "")).lower()


def format_mac_colon(mac_norm):
    if len(mac_norm) != 12:
        return mac_norm
    return ":".join(mac_norm[i:i + 2] for i in range(0, 12, 2))


def first_mac_in_text(text):
    patterns = [
        r"(?:[0-9a-fA-F]{2}[:.-]){5}[0-9a-fA-F]{2}",
        r"(?:[0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4}",
    ]
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return m.group(0)
    return "Unknown"


def first_ip_in_text(text):
    m = re.search(r"\b(?:\d{1,3}\.){3}\d{1,3}\b", text)
    return m.group(0) if m else "Unknown"


def get_lock_owner(path):
    lock_path = path.parent / f"~${path.name}"
    if not lock_path.exists():
        return "Unknown (Closed)"
    try:
        content = lock_path.read_bytes().decode("latin-1", errors="ignore")
        m = re.search(r"[a-zA-Z\s]{3,}", content)
        return m.group(0).strip() if m else "a Colleague"
    except Exception:
        return "a Colleague"


def is_file_locked(path):
    try:
        fd = os.open(path, os.O_RDWR | os.O_APPEND)
        os.close(fd)
        return False
    except OSError as exc:
        return exc.errno in (errno.EACCES, errno.EPERM, errno.ETXTBSY)


def is_workbook_open(path):
    lock_path = path.parent / f"~${path.name}"
    if lock_path.exists():
        return True
    return is_file_locked(path)


def confirm_change(safe_mode, switch_ip, port, current_vlan, target_vlan, row_idx, workbook_name):
    if not safe_mode:
        return True

    if not os.isatty(0):
        raise RuntimeError("--safe was supplied, but no interactive terminal is available.")

    print("\n" + "=" * 60)
    print(f"[ACTION REQUIRED] File: {workbook_name}")
    print(f"Row {row_idx}: {switch_ip} {port}")
    print(f"Live VLAN: [{current_vlan}] -> Target VLAN: [{target_vlan}]")
    return input("Apply change? (y/n): ").strip().lower() == "y"


def resolve_workbook_paths():
    resolved = []
    for filename in FILES_TO_CHECK:
        found = None
        for base_dir in SEARCH_DIRS:
            candidate = base_dir / filename
            if candidate.exists():
                found = candidate
                break
        if found:
            resolved.append(found)
        else:
            resolved.append(SEARCH_DIRS[0] / filename)
    return resolved


# ============================================================
# Teams
# ============================================================
def send_teams_notification(status, message, workbook_path, details=None, last_edited_by=None):
    if not TEAMS_WEBHOOK_URL:
        return

    color = {
        "SUCCESS": "28A745",
        "WARNING": "FFC107",
        "CRITICAL": "DC3545",
    }.get(status, "0078D7")

    facts = [
        {"name": "File", "value": workbook_path.name},
    ]

    if last_edited_by:
        facts.append({"name": "Last Edited", "value": f"{last_edited_by} @ {now_friendly()}"})

    if details:
        for entry in details:
            location_str = entry.get("location", "n/a")
            entry_status = entry.get("status", "Updated")
            entry_notes = entry.get("notes", "n/a") or "n/a"
            facts.append(
                {
                    "name": f"Switch {entry.get('switch_ip', 'Unknown')}",
                    "value": (
                        f"Port {entry['port']} - {location_str} - {entry_status} {entry['target_vlan']} --- "
                        f"previously {entry.get('old_vlan', 'Unknown')} MAC: {entry.get('mac', 'Unknown')} "
                        f"IP: {entry.get('device_ip', 'Unknown')} -- Notes : {entry_notes}"
                    ),
                }
            )

    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": color,
        "summary": "Aruba Config Update",
        "sections": [
            {
                "activityTitle": f"**Aruba Configurator: {status}**",
                "activitySubtitle": f"File: {workbook_path.name}",
                "text": message,
                "facts": facts,
            }
        ],
    }

    try:
        r = requests.post(TEAMS_WEBHOOK_URL, json=payload, timeout=10)
        r.raise_for_status()
    except Exception as exc:
        log(f"[!] Teams Alert Failed for {workbook_path.name}: {exc}")


# ============================================================
# Excel
# ============================================================
def build_header_map(ws):
    aliases = {
        "target_vlan": {"vlan"},
        "switch_ip": {"switch ip"},
        "port": {"port"},
        "mac": {"mac"},
        "ip": {"ip"},
        "last_checked": {"last checked"},
        "notes": {"notes"},
        "room": {"room", "cr"},
        "outlet": {"outlet"},
        "building": {"building", "bldg"},
    }

    header_map = {}
    for col in range(1, ws.max_column + 1):
        for row in (2, 3):
            header = normalize_header(ws.cell(row=row, column=col).value)
            if not header:
                continue
            for key, names in aliases.items():
                if header in names:
                    header_map[key] = col
    return header_map


def read_cell(ws, row_idx, col_idx):
    return clean_text(ws.cell(row=row_idx, column=col_idx).value)


def write_readonly_columns(ws, row_idx, header_map, mac=None, ip=None, last_checked=None, notes=None):
    if mac is not None and "mac" in header_map:
        ws.cell(row=row_idx, column=header_map["mac"], value=mac)
    if ip is not None and "ip" in header_map:
        ws.cell(row=row_idx, column=header_map["ip"], value=ip)
    if last_checked is not None and "last_checked" in header_map:
        ws.cell(row=row_idx, column=header_map["last_checked"], value=last_checked)
    if notes is not None and "notes" in header_map:
        ws.cell(row=row_idx, column=header_map["notes"], value=notes)


def write_switch_failure_row(ws, row_idx, header_map, checked_at):
    existing_mac = read_cell(ws, row_idx, header_map["mac"])
    existing_ip = read_cell(ws, row_idx, header_map["ip"])

    mac_value = existing_mac if existing_mac else "Error"
    ip_value = existing_ip if existing_ip else "Error"

    write_readonly_columns(
        ws,
        row_idx,
        header_map,
        mac=mac_value,
        ip=ip_value,
        last_checked=checked_at,
        notes="Switch connection failed",
    )


# ============================================================
# Network helpers
# ============================================================
def connect_device(host, device_type="aruba_aoscx"):
    return ConnectHandler(
        device_type=device_type,
        host=host,
        username=USERNAME,
        password=PASSWORD,
        conn_timeout=20,
        fast_cli=False,
    )


def prepare_session(net_connect):
    for cmd in ("no page", "aruba-central support-mode"):
        try:
            net_connect.send_command_timing(cmd)
        except Exception:
            pass


def parse_show_int_br(output):
    port_vlan_map = {}
    for line in output.splitlines():
        m = re.match(r"^\s*(\d+/\d+/\d+)\s+(\S+)", line.rstrip())
        if m:
            port_vlan_map[m.group(1).strip()] = m.group(2).strip()
    return port_vlan_map


def get_live_vlan_map(net_connect):
    return parse_show_int_br(net_connect.send_command("show int br", read_timeout=60))


def get_port_mac(net_connect, port):
    for cmd in (
        f"show mac-address-table interface {port}",
        f"show mac-address-table int {port}",
    ):
        try:
            output = net_connect.send_command(cmd, read_timeout=30)
            mac = first_mac_in_text(output)
            if mac != "Unknown":
                return mac
        except Exception:
            pass

    try:
        output = net_connect.send_command("show mac-address-table", read_timeout=60)
        for line in output.splitlines():
            if port in line:
                mac = first_mac_in_text(line)
                if mac != "Unknown":
                    return mac
    except Exception:
        pass

    return "Unknown"


def build_router_arp_cache(router_connect):
    arp_cache = {}

    for cmd in ("show arp", "show arp all-vrfs"):
        try:
            output = router_connect.send_command(cmd, read_timeout=120)
        except Exception:
            continue

        for line in output.splitlines():
            ip = first_ip_in_text(line)
            mac = first_mac_in_text(line)
            if ip == "Unknown" or mac == "Unknown":
                continue

            mac_norm = normalize_mac(mac)
            if mac_norm and mac_norm not in arp_cache:
                arp_cache[mac_norm] = ip

    return arp_cache


def get_ip_for_mac(router_connect, arp_cache, mac):
    if mac == "Unknown":
        return "Unknown"

    mac_norm = normalize_mac(mac)
    if not mac_norm:
        return "Unknown"

    if mac_norm in arp_cache:
        return arp_cache[mac_norm]

    if not router_connect:
        return "Unknown"

    search_terms = {
        mac,
        format_mac_colon(mac_norm),
        f"{mac_norm[0:4]}.{mac_norm[4:8]}.{mac_norm[8:12]}",
    }

    for term in search_terms:
        for cmd in (
            f"show arp | include {term}",
            f"show arp | inc {term}",
            f"show arp all-vrfs | include {term}",
            f"show arp all-vrfs | inc {term}",
        ):
            try:
                output = router_connect.send_command(cmd, read_timeout=60)
            except Exception:
                continue

            for line in output.splitlines():
                if mac_norm in normalize_mac(line):
                    ip = first_ip_in_text(line)
                    if ip != "Unknown":
                        return ip

    return "Unknown"


def get_port_live_details(net_connect, port, router_connect, arp_cache):
    mac = get_port_mac(net_connect, port)
    ip = get_ip_for_mac(router_connect, arp_cache, mac)
    return {"mac": mac, "ip": ip}


def apply_vlan_change(net_connect, port, target_vlan):
    output = ""
    output += net_connect.send_command_timing("configure terminal")
    output += net_connect.send_command_timing(f"interface {port}")
    output += net_connect.send_command_timing(f"vlan access {target_vlan}")
    output += net_connect.send_command_timing("end")
    return output


# ============================================================
# Workbook processing
# ============================================================
def process_workbook(workbook_path, args, router_connect, arp_cache):
    result = {
        "checked": 0,
        "candidates": 0,
        "changed": 0,
        "already_correct": 0,
        "declined": 0,
        "failed": 0,
        "processed": False,
        "skipped": False,
    }

    log("")
    log("=" * 80)
    log(f"[*] Processing workbook: {workbook_path}")

    if not workbook_path.exists():
        log(f"[!] File not found, skipping: {workbook_path}")
        result["skipped"] = True
        return result

    start_mtime = os.path.getmtime(workbook_path)

    try:
        wb = load_workbook(workbook_path, data_only=False)
        ws = wb.active
    except Exception as exc:
        log(f"[!] Error loading workbook {workbook_path.name}: {exc}")
        result["failed"] += 1
        return result

    header_map = build_header_map(ws)
    required = ("target_vlan", "switch_ip", "port", "mac", "ip", "last_checked", "notes")
    missing = [key for key in required if key not in header_map]
    if missing:
        log(f"[!] ERROR in {workbook_path.name}: Missing required headers: {missing}")
        result["failed"] += 1
        return result

    last_edited_by = getattr(wb.properties, "lastModifiedBy", None) or "Unknown"

    # Optional location headers
    has_location = all(k in header_map for k in ("building", "room", "outlet"))

    rows_by_switch = defaultdict(list)

    for row_idx in range(4, ws.max_row + 1):
        switch_ip = clean_text(ws.cell(row=row_idx, column=header_map["switch_ip"]).value)
        port = clean_text(ws.cell(row=row_idx, column=header_map["port"]).value)
        target_vlan = clean_text(ws.cell(row=row_idx, column=header_map["target_vlan"]).value)

        if switch_ip and port and target_vlan:
            entry = {
                "row_idx": row_idx,
                "port": port,
                "target_vlan": target_vlan,
                "notes": clean_text(ws.cell(row=row_idx, column=header_map["notes"]).value),
            }
            
            # Extract location if available
            if has_location:
                building = clean_text(ws.cell(row=row_idx, column=header_map["building"]).value)
                room = clean_text(ws.cell(row=row_idx, column=header_map["room"]).value)
                outlet = clean_text(ws.cell(row=row_idx, column=header_map["outlet"]).value)
                entry["location"] = f"{building}/{room}/{outlet}"
            
            rows_by_switch[switch_ip].append(entry)

    if not rows_by_switch:
        log(f"[*] No usable rows found in {workbook_path.name}")
        result["processed"] = True
        return result

    workbook_touched = False
    report_entries = []

    for switch_ip, row_entries in rows_by_switch.items():
        log(f"[*] {workbook_path.name} | Connecting to switch {switch_ip} for {len(row_entries)} row(s)")

        try:
            with connect_device(switch_ip, "aruba_aoscx") as net_connect:
                prepare_session(net_connect)

                live_vlan_map_before = get_live_vlan_map(net_connect)
                pending_verification = []

                for entry in row_entries:
                    row_idx = entry["row_idx"]
                    port = entry["port"]
                    target_vlan = entry["target_vlan"]
                    checked_at = now_str()

                    current_vlan = live_vlan_map_before.get(port, "Unknown")
                    live_details = get_port_live_details(net_connect, port, router_connect, arp_cache)

                    result["checked"] += 1

                    if current_vlan == target_vlan:
                        result["already_correct"] += 1
                        location_str = f" | Outlet: {entry['location']}" if "location" in entry else ""
                        log(
                            f"[OK] {workbook_path.name} | Row {row_idx} | {switch_ip} | Port {port}{location_str} "
                            f"already on VLAN {target_vlan} | MAC: {live_details['mac']} | IP: {live_details['ip']}"
                        )
                        if not args.dry_run:
                            write_readonly_columns(
                                ws,
                                row_idx,
                                header_map,
                                mac=live_details["mac"],
                                ip=live_details["ip"],
                                last_checked=checked_at,
                                notes="",
                            )
                            workbook_touched = True
                        continue

                    result["candidates"] += 1

                    if args.dry_run:
                        location_str = f" | Outlet: {entry['location']}" if "location" in entry else ""
                        log(
                            f"[DRY-RUN] {workbook_path.name} | Row {row_idx} | {switch_ip} | Port {port}{location_str} | "
                            f"Live VLAN {current_vlan} -> Target VLAN {target_vlan} | "
                            f"MAC: {live_details['mac']} | IP: {live_details['ip']}"
                        )
                        continue

                    should_apply = confirm_change(
                        safe_mode=args.safe,
                        switch_ip=switch_ip,
                        port=port,
                        current_vlan=current_vlan,
                        target_vlan=target_vlan,
                        row_idx=row_idx,
                        workbook_name=workbook_path.name,
                    )

                    if not should_apply:
                        result["declined"] += 1
                        location_str = f" | Outlet: {entry['location']}" if "location" in entry else ""
                        log(f"[SKIPPED] {workbook_path.name} | Row {row_idx}{location_str} declined by user.")
                        report_entries.append(
                            {
                                "switch_ip": switch_ip,
                                "port": port,
                                "location": entry.get("location", "n/a"),
                                "target_vlan": target_vlan,
                                "old_vlan": current_vlan,
                                "mac": live_details["mac"],
                                "device_ip": live_details["ip"],
                                "notes": entry.get("notes", "n/a") or "n/a",
                                "status": "denied Vlan",
                            }
                        )
                        write_readonly_columns(
                            ws,
                            row_idx,
                            header_map,
                            mac=live_details["mac"],
                            ip=live_details["ip"],
                            last_checked=checked_at,
                            notes=f"Current VLAN: {current_vlan}",
                        )
                        workbook_touched = True
                        continue

                    changed_by = RUN_ACTOR
                    changed_at = now_str()
                    source = RUN_SOURCE

                    location_str = f" | Outlet: {entry['location']}" if "location" in entry else ""
                    log(
                        f"[*] Applying {workbook_path.name} | Row {row_idx} | Switch {switch_ip} | Port {port}{location_str} | "
                        f"{current_vlan} -> {target_vlan} | By: {changed_by} | At: {changed_at} | Source: {source}"
                    )

                    try:
                        apply_vlan_change(net_connect, port, target_vlan)
                        verify_item = {
                            "row_idx": row_idx,
                            "port": port,
                            "target_vlan": target_vlan,
                            "old_vlan": current_vlan,
                            "notes": entry.get("notes", "n/a") or "n/a",
                            "changed_by": changed_by,
                            "changed_at": changed_at,
                            "source": source,
                        }
                        if "location" in entry:
                            verify_item["location"] = entry["location"]
                        pending_verification.append(verify_item)
                    except Exception as exc:
                        result["failed"] += 1
                        location_str = f" | Outlet: {entry['location']}" if "location" in entry else ""
                        log(
                            f"[FAILED] {workbook_path.name} | Row {row_idx} | {switch_ip} | Port {port}{location_str} | "
                            f"Error during config: {str(exc)[:200]}"
                        )
                        report_entries.append(
                            {
                                "switch_ip": switch_ip,
                                "port": port,
                                "location": entry.get("location", "n/a"),
                                "target_vlan": target_vlan,
                                "old_vlan": current_vlan,
                                "mac": live_details["mac"],
                                "device_ip": live_details["ip"],
                                "notes": entry.get("notes", "n/a") or "n/a",
                                "status": "failed Vlan",
                            }
                        )
                        write_readonly_columns(
                            ws,
                            row_idx,
                            header_map,
                            mac=live_details["mac"],
                            ip=live_details["ip"],
                            last_checked=checked_at,
                            notes=f"Current VLAN: {current_vlan}",
                        )
                        workbook_touched = True

                if pending_verification:
                    live_vlan_map_after = get_live_vlan_map(net_connect)

                    for item in pending_verification:
                        row_idx = item["row_idx"]
                        port = item["port"]
                        target_vlan = item["target_vlan"]
                        old_vlan = item["old_vlan"]
                        verified_vlan = live_vlan_map_after.get(port, "Unknown")
                        checked_at = now_str()
                        live_details = get_port_live_details(net_connect, port, router_connect, arp_cache)

                        if verified_vlan == target_vlan:
                            result["changed"] += 1
                            location_str = f" | Outlet: {item['location']}" if "location" in item else ""
                            log(
                                f"[DONE] {workbook_path.name} | Row {row_idx} | {switch_ip} | Port {port}{location_str} "
                                f"-> VLAN {target_vlan} | MAC: {live_details['mac']} | IP: {live_details['ip']} | "
                                f"By: {item['changed_by']} | At: {item['changed_at']} | Source: {item['source']}"
                            )
                            write_readonly_columns(
                                ws,
                                row_idx,
                                header_map,
                                mac=live_details["mac"],
                                ip=live_details["ip"],
                                last_checked=checked_at,
                                notes="",
                            )
                            workbook_touched = True

                            summary_entry = {
                                "switch_ip": switch_ip,
                                "port": port,
                                "location": item.get("location", "n/a"),
                                "target_vlan": target_vlan,
                                "old_vlan": old_vlan,
                                "mac": live_details["mac"],
                                "device_ip": live_details["ip"],
                                "notes": item.get("notes", "n/a") or "n/a",
                                "status": "New Vlan",
                                "changed_by": item["changed_by"],
                                "changed_at": item["changed_at"],
                                "source": item["source"],
                            }
                            if "location" in item:
                                summary_entry["location"] = item["location"]
                            report_entries.append(summary_entry)
                        else:
                            result["failed"] += 1
                            location_str = f" | Outlet: {item['location']}" if "location" in item else ""
                            log(
                                f"[FAILED] {workbook_path.name} | Row {row_idx} | {switch_ip} | Port {port}{location_str} | "
                                f"Target VLAN {target_vlan} not applied. Live VLAN is still {verified_vlan}"
                            )
                            report_entries.append(
                                {
                                    "switch_ip": switch_ip,
                                    "port": port,
                                    "location": item.get("location", "n/a"),
                                    "target_vlan": target_vlan,
                                    "old_vlan": old_vlan,
                                    "mac": live_details["mac"],
                                    "device_ip": live_details["ip"],
                                    "notes": item.get("notes", "n/a") or "n/a",
                                    "status": "failed Vlan",
                                }
                            )
                            write_readonly_columns(
                                ws,
                                row_idx,
                                header_map,
                                mac=live_details["mac"],
                                ip=live_details["ip"],
                                last_checked=checked_at,
                                notes=f"Current VLAN: {verified_vlan}",
                            )
                            workbook_touched = True

        except Exception as exc:
            result["failed"] += len(row_entries)
            log(f"[!] Switch-level failure in {workbook_path.name} on {switch_ip}: {str(exc)[:250]}")

            if not args.dry_run:
                checked_at = now_str()
                for entry in row_entries:
                    write_switch_failure_row(ws, entry["row_idx"], header_map, checked_at)
                workbook_touched = True

    if args.dry_run:
        result["processed"] = True
        log(
            f"[*] Dry-run summary for {workbook_path.name} | "
            f"Checked: {result['checked']} | Candidates: {result['candidates']} | "
            f"Changed: {result['changed']} | Already Correct: {result['already_correct']} | "
            f"Declined: {result['declined']} | Failed: {result['failed']}"
        )
        return result

    if workbook_touched:
        if os.path.getmtime(workbook_path) != start_mtime:
            owner = get_lock_owner(workbook_path)
            send_teams_notification(
                "CRITICAL",
                f"Conflict detected. {owner} modified the file while the script was running.",
                workbook_path,
            )
            log(f"[!] File changed while script was running. Save aborted for {workbook_path.name}")
            result["failed"] += 1
            return result

        try:
            wb.save(workbook_path)
            log(f"[+] Success: Spreadsheet updated: {workbook_path.name}")
        except PermissionError:
            owner = get_lock_owner(workbook_path)
            log(f"[!] SAVE FAILED for {workbook_path.name}: {owner} has the file open.")
            result["failed"] += 1
            return result
        except Exception as exc:
            log(f"[!] SAVE FAILED for {workbook_path.name}: {exc}")
            result["failed"] += 1
            return result

    if result["failed"] > 0:
        send_teams_notification(
            "WARNING",
            f"Completed with errors. Changed {result['changed']} port(s), but {result['failed']} row(s) failed.",
            workbook_path,
            details=report_entries,
            last_edited_by=last_edited_by,
        )
    elif result["changed"] > 0:
        send_teams_notification(
            "SUCCESS",
            f"Successfully updated {result['changed']} port(s).",
            workbook_path,
            details=report_entries,
            last_edited_by=last_edited_by,
        )

    result["processed"] = True
    return result


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--safe", action="store_true", help="Confirm each VLAN change manually")
    parser.add_argument("--dry-run", action="store_true", help="Show what would change without applying it")
    args = parser.parse_args()

    log("[*] Script starting")
    log(f"[*] Safe mode: {args.safe}")
    log(f"[*] Dry run: {args.dry_run}")
    log(f"[*] Run By: {RUN_ACTOR}")
    log(f"[*] Source: {RUN_SOURCE}")
    log(f"[*] Router ARP lookup target: {ROUTER_IP}")

    if not USERNAME or not PASSWORD:
        log("[!] ERROR: Missing username/passwordAD environment variables.")
        sys.exit(1)

    workbook_paths = resolve_workbook_paths()

    locked_files = [path for path in workbook_paths if is_workbook_open(path)]
    if locked_files:
        log("[!] Aborting: the following workbook(s) appear to be open and may conflict:")
        details = []
        for path in locked_files:
            owner = get_lock_owner(path)
            log(f"    - {path.name} (locked by {owner})")
            details.append(
                {
                    "switch_ip": "N/A",
                    "port": "N/A",
                    "location": "N/A",
                    "target_vlan": "N/A",
                    "old_vlan": "N/A",
                    "mac": "N/A",
                    "device_ip": "N/A",
                    "notes": f"Locked by {owner}",
                    "status": "file locked",
                }
            )

        send_teams_notification(
            "CRITICAL",
            "Script aborted because one or more Excel files are open and unavailable.",
            locked_files[0],
            details=details,
        )
        sys.exit(1)

    log("[*] Files queued for checking:")
    for path in workbook_paths:
        log(f"    - {path}")

    if not workbook_paths:
        log("[!] No files configured in FILES_TO_CHECK.")
        sys.exit(1)

    router_connect = None
    arp_cache = {}

    try:
        log(f"[*] Connecting to router for ARP lookups: {ROUTER_IP}")
        router_connect = connect_device(ROUTER_IP, ROUTER_DEVICE_TYPE)
        prepare_session(router_connect)
        arp_cache = build_router_arp_cache(router_connect)
        log(f"[*] Router ARP cache loaded: {len(arp_cache)} MAC/IP entries")
    except Exception as exc:
        log(f"[!] Router ARP lookup unavailable ({ROUTER_IP}): {str(exc)[:250]}")
        router_connect = None
        arp_cache = {}

    overall = {
        "checked": 0,
        "candidates": 0,
        "changed": 0,
        "already_correct": 0,
        "declined": 0,
        "failed": 0,
        "processed_files": 0,
        "skipped_files": 0,
    }

    try:
        for workbook_path in workbook_paths:
            if is_workbook_open(workbook_path):
                owner = get_lock_owner(workbook_path)
                log(f"[!] Aborting: {workbook_path.name} is currently open by {owner}.")
                sys.exit(1)

            result = process_workbook(workbook_path, args, router_connect, arp_cache)

            overall["checked"] += result["checked"]
            overall["candidates"] += result["candidates"]
            overall["changed"] += result["changed"]
            overall["already_correct"] += result["already_correct"]
            overall["declined"] += result["declined"]
            overall["failed"] += result["failed"]

            if result["processed"]:
                overall["processed_files"] += 1
            if result["skipped"]:
                overall["skipped_files"] += 1

    finally:
        if router_connect:
            try:
                router_connect.disconnect()
            except Exception:
                pass

    log("")
    log("=" * 80)
    log(
        f"[*] Overall Summary | Files processed: {overall['processed_files']} | "
        f"Files skipped: {overall['skipped_files']} | Checked: {overall['checked']} | "
        f"Candidates: {overall['candidates']} | Changed: {overall['changed']} | "
        f"Already Correct: {overall['already_correct']} | Declined: {overall['declined']} | "
        f"Failed: {overall['failed']}"
    )

    if args.dry_run:
        log("[*] Dry run finished successfully.")
        sys.exit(0)

    if overall["failed"] > 0:
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()