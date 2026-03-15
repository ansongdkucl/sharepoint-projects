import os
import sys
import time
import msal
import requests
import logging
from io import BytesIO
from openpyxl import load_workbook
from netmiko import ConnectHandler
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv

# ============================================================
# 1. CONFIGURATION (Graph API & SSH)
# ============================================================
load_dotenv()
CLIENT_ID = "83591eda-f950-42b7-b09b-d2e8ad59305c"
TENANT_ID = "1faf88fe-a998-4c5b-93c9-210a11d9a5c2"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["Sites.Selected", "User.Read"]
DRIVE_ID = "b!ICFlXjjf50u1m5wQqA2CV48VyCYy_4VAqI_kt0Xcz9pXIF50OxppT50xCaaoRbQa"
ITEM_ID  = "01V4ZDPCOA6S7N23QEVREKL5YF2QZZXW7U"

USERNAME = os.getenv("username")
PASSWORD = os.getenv("passwordAD")
MAX_SWITCH_CONCURRENCY = 3 # Safe limit for session stability

# ============================================================
# 2. SHAREPOINT AUTH & DOWNLOAD
# ============================================================
def get_access_token():
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result: return result.get("access_token")
    
    flow = app.initiate_device_flow(scopes=SCOPES)
    print(f"\n{flow['message']}\n")
    result = app.acquire_token_by_device_flow(flow)
    return result.get("access_token")

# ============================================================
# 3. NETWORK LOGIC (Fast & Logged)
# ============================================================
def process_switch_queue(switch_ip, tasks):
    device = {
        'device_type': 'aruba_osswitch',
        'host': switch_ip,
        'username': USERNAME,
        'password': PASSWORD,
        'global_delay_factor': 2.5, # Critical for "Pattern Not Detected" errors
        'session_log': f'log_{switch_ip}.txt' 
    }
    
    results = []
    try:
        with ConnectHandler(**device) as net_connect:
            # Handle Aruba Support Mode
            out = net_connect.send_command("aruba-central support-mode", expect_string=r"y/n|#")
            if "y/n" in out:
                net_connect.send_command("y", expect_string=r"#")
            
            net_connect.send_command("conf t", expect_string=r"\(config\)#")

            for row_idx, port, description in tasks:
                print(f"    [~] {switch_ip} | Row {row_idx} | Updating Port {port}")
                net_connect.send_command(f"interface {port}", expect_string=r"\(config-if\)#")
                net_connect.send_command(f"description \"{description}\"", expect_string=r"\(config-if\)#")
                net_connect.send_command("exit", expect_string=r"\(config\)#")
                results.append(f"[ROW {row_idx}] SUCCESS: {switch_ip} {port}")
            
            net_connect.send_command("write memory", expect_string=r"#")
    except Exception as e:
        results.append(f"[ERROR] {switch_ip}: {str(e)[:50]}")
    
    return results

# ============================================================
# 4. MAIN WORKFLOW
# ============================================================
def main():
    token = get_access_token()
    if not token: return

    # Download File from SharePoint
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers)
    
    if r.status_code != 200:
        print(f"Failed to fetch file: {r.status_code}")
        return

    # Load Excel from Memory
    wb = load_workbook(BytesIO(r.content), data_only=True)
    ws = wb.active
    
    # Batch switches
    switch_batches = {}
    for row in range(4, ws.max_row + 1):
        sw_ip = ws.cell(row=row, column=8).value
        port  = ws.cell(row=row, column=9).value
        desc  = f"{ws.cell(row=row, column=1).value}/{ws.cell(row=row, column=2).value}/{ws.cell(row=row_idx, column=3).value}"
        
        if sw_ip and port:
            ip = str(sw_ip).strip()
            if ip not in switch_batches: switch_batches[ip] = []
            switch_batches[ip].append((row, str(port).strip(), desc))

    # Parallel Execution
    print(f"[*] Starting update on {len(switch_batches)} switches...")
    with ThreadPoolExecutor(max_workers=MAX_SWITCH_CONCURRENCY) as executor:
        futures = [executor.submit(process_switch_queue, ip, tasks) for ip, tasks in switch_batches.items()]
        for f in futures:
            for res in f.result(): print(res)

if __name__ == "__main__":
    main()