import os
import re
import sys
import argparse
import time
import datetime
import requests
import json
from pathlib import Path
from openpyxl import load_workbook
from netmiko import ConnectHandler
from dotenv import load_dotenv

# ============================================================
# 1. SETUP & CONFIGURATION
# ============================================================
load_dotenv()

USERNAME = os.getenv("username")
PASSWORD = os.getenv("passwordAD")

# Webhook URL
TEAMS_WEBHOOK_URL = "https://liveuclac.webhook.office.com/webhookb2/4bf41ac1-0d61-4760-9dea-e0f6184dde8a@1faf88fe-a998-4c5b-93c9-210a11d9a5c2/IncomingWebhook/46bcd5e8d47e4de5b575fe10f189b1e1/43bfe760-7689-4d0b-96fd-46b265519580/V2OHjE1yNbt-Vi6fJNAiAdSwjM90ZnyOKY49V-zdJi0dA1"

# Path to the file on N: drive
DEFAULT_PATH = Path("N:/Patching Schedule/90 TCR - Level 3A Patching Schedule.xlsx")
# If testing locally on WSL, uncomment the line below:
# DEFAULT_PATH = Path("/mnt/c/Users/cceadan/OneDrive - University College London/Estates IT - Project Documentation - Patching Schedule/90TCR - Daniel Test.xlsx")

LOG_FILE = DEFAULT_PATH.parent / "patching_audit_log.txt"

def log_event(message):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {message}\n"
    with open(LOG_FILE, "a") as f:
        f.write(log_entry)
    print(f"[*] Logged: {message}")

def send_teams_notification(status, message, details=None):
    if not TEAMS_WEBHOOK_URL or "placeholder" in TEAMS_WEBHOOK_URL:
        return 
    color = {"SUCCESS": "28A745", "WARNING": "FFC107", "CRITICAL": "DC3545"}.get(status, "0078D7")
    facts = []
    if details:
        for entry in details:
            facts.append({"name": f"Switch {entry['ip']}", "value": f"Port {entry['port']} -> VLAN {entry['vlan']}"})
    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": color,
        "summary": "Aruba Config Update",
        "sections": [{
            "activityTitle": f"**Aruba Configurator: {status}**",
            "activitySubtitle": f"File: {DEFAULT_PATH.name}",
            "text": message,
            "facts": facts
        }]
    }
    try:
        requests.post(TEAMS_WEBHOOK_URL, json=payload, timeout=10)
    except Exception as e:
        print(f"[!] Teams Alert Failed: {e}")

def get_lock_owner(path):
    lock_path = path.parent / f"~${path.name}"
    if not lock_path.exists(): return "Unknown (Closed)"
    try:
        with open(lock_path, 'rb') as f:
            content = f.read().decode('latin-1', errors='ignore')
            match = re.search(r'[a-zA-Z\s]{3,}', content)
            return match.group(0).strip() if match else "a Colleague"
    except: return "a Colleague"

def run_aruba_config(switch_ip, port, vlan_id):
    device = {'device_type': 'aruba_aoscx', 'host': switch_ip, 'username': USERNAME, 'password': PASSWORD}
    commands = ["aruba-central support-mode", "conf t", f"int {port}", f"vlan access {vlan_id}"]
    try:
        with ConnectHandler(**device, conn_timeout=15) as net_connect:
            net_connect.send_config_set(commands)
            mac_out = net_connect.send_command(f"show mac-address-table int {port}")
            mac_match = re.search(r'([0-9a-fA-F]{2}[:.-]){5}[0-9a-fA-F]{2}', mac_out)
            found_mac = mac_match.group(0) if mac_match else "Unknown"
            arp_out = net_connect.send_command(f"show arp | inc {found_mac}")
            ip_match = re.search(r'(\d{1,3}\.){3}\d{1,3}', arp_out)
            found_ip = ip_match.group(0) if ip_match else "No ARP"
            return {"mac": found_mac, "ip": found_ip, "status": "Success"}
    except Exception as e:
        log_event(f"FAILURE: {switch_ip} Port {port} - Error: {str(e)}")
        return {"mac": "Error", "ip": "Error", "status": str(e)[:20]}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--safe", action="store_true", help="Confirm each change manually")
    parser.add_argument("--dry-run", action="store_true", help="Show changes without applying them")
    args = parser.parse_args()

    if not DEFAULT_PATH.exists(): 
        print(f"[!] File not found: {DEFAULT_PATH}")
        return

    if (DEFAULT_PATH.parent / f"~${DEFAULT_PATH.name}").exists():
        owner = get_lock_owner(DEFAULT_PATH)
        print(f"\n[!] ABORTED: File is currently open by {owner}.")
        return

    try:
        wb = load_workbook(DEFAULT_PATH, data_only=False)
        ws = wb.active
        print(f"[*] File Loaded: {DEFAULT_PATH.name}")
    except Exception as e:
        print(f"Error Loading Workbook: {e}"); return

    # 2. Header Mapping
    read_map = {}   
    write_map = {}  
    
    for col in range(1, ws.max_column + 1):
        v2 = str(ws.cell(row=2, column=col).value or "").strip()
        v3 = str(ws.cell(row=3, column=col).value or "").strip()
        
        for h in [v2, v3]:
            if h == "VLAN":      read_map["vlan"] = col
            if h == "SWITCH IP": read_map["switch"] = col
            if h == "PORT":      read_map["port"] = col
            if h == "vlan":      write_map["vlan"] = col
            if h == "switch":    write_map["switch"] = col
            if h == "port":      write_map["port"] = col
            if h == "mac":       write_map["mac"] = col
            if h == "ip":        write_map["ip"] = col

    print(f"\n--- Column Mapping Check ---")
    print(f"Target VLAN (G): Col {read_map.get('vlan')} | Status vlan (O): Col {write_map.get('vlan')}")
    
    if not read_map.get("vlan") or not write_map.get("vlan"):
        print("[!] ERROR: Header mapping failed. Check Row 2/3 for VLAN vs vlan.")
        return

    # 3. Processing
    rows_processed = 0
    rows_skipped = 0
    config_summary = []

    for row_idx in range(4, ws.max_row + 1):
        v_target = ws.cell(row=row_idx, column=read_map["vlan"]).value
        v_current = ws.cell(row=row_idx, column=write_map["vlan"]).value
        s_val = ws.cell(row=row_idx, column=read_map["switch"]).value
        p_val = ws.cell(row=row_idx, column=read_map["port"]).value

        # Skip rows missing core switch info
        if not s_val or not p_val or v_target is None:
            continue

        # Port formatting (Handles 1/1/x becoming dates)
        if hasattr(p_val, 'strftime'):
            p_val = f"1/1/{p_val.day}" 
        
        p_val = str(p_val).strip()
        target_str = str(v_target).strip()
        current_str = str(v_current).strip() if v_current is not None else "Empty/New"

        # --- PRE-CHECK DEBUG ---
        # print(f"Row {row_idx}: Mapping Check [{current_str}] -> [{target_str}]")

        # Skip if they match
        if target_str == current_str:
            rows_skipped += 1
            continue

        if args.dry_run:
            print(f"[DRY-RUN] Row {row_idx}: {s_val} | Port {p_val} | Current: {current_str} -> Target: {target_str}")
            rows_processed += 1
            continue

        # --- USER PROMPT ---
        if args.safe:
            print(f"\n" + "-"*40)
            print(f"PROPOSED CHANGE: Row {row_idx}")
            print(f"Switch: {s_val} | Port: {p_val}")
            print(f"VLAN:   [Current: {current_str}] ---> [Target: {target_str}]")
            if input("Apply Config? (y/n): ").lower() != 'y': 
                print("[*] Skipped by user.")
                continue

        # EXECUTION
        res = run_aruba_config(str(s_val), p_val, target_str)
        
        if res["status"] == "Success":
            log_event(f"SUCCESS: Row {row_idx} | {s_val} Port {p_val} | {current_str} -> {target_str}")
            config_summary.append({"ip": s_val, "port": p_val, "vlan": target_str})
            
            # Update Excel status columns (Column O and neighbors)
            ws.cell(row=row_idx, column=write_map["switch"], value=s_val)
            ws.cell(row=row_idx, column=write_map["port"], value=p_val)
            ws.cell(row=row_idx, column=write_map["vlan"], value=v_target)
            ws.cell(row=row_idx, column=write_map["mac"], value=res["mac"])
            ws.cell(row=row_idx, column=write_map["ip"], value=res["ip"])
            rows_processed += 1
            print(f"    [DONE] Row {row_idx} updated.")

    # 4. Finalize
    if rows_processed > 0 and not args.dry_run:
        try:
            wb.save(DEFAULT_PATH)
            send_teams_notification("SUCCESS", f"Updated {rows_processed} ports on {DEFAULT_PATH.name}.", details=config_summary)
            print(f"\n[+] Success: {rows_processed} changes saved to Excel.")
        except PermissionError:
            log_event("CRITICAL: Save failed - File locked.")
            print(f"\n[!] SAVE FAILED: File is open.")
    else:
        print(f"\n[*] Run Complete. Updated: {rows_processed} | Already Correct: {rows_skipped}")

if __name__ == "__main__":
    main()