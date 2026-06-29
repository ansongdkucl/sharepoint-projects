import os, re, sys, argparse, requests, platform, tempfile, subprocess, time
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from netmiko import ConnectHandler
from dotenv import load_dotenv

# ============================================================
# 1. SETUP, VALUES & CONFIGURATION
# ============================================================
load_dotenv()

USERNAME = os.getenv("username")
PASSWORD = os.getenv("passwordAD")
TEAMS_WEBHOOK_URL = os.getenv("TEAMS_WEBHOOK_URL", "")

RUN_ACTOR = os.getenv("GITHUB_ACTOR") or os.getenv("USER") or os.getenv("USERNAME") or "unknown"
RUN_SOURCE = "GitHub Actions" if os.getenv("GITHUB_ACTIONS", "").lower() == "true" else "Manual"
WORKBOOK_NAME = "FC-MSA-CI.xlsx"
#WORKBOOK_NAME = "FC-MSA-CI-test1.xlsx"
WORKBOOK_DIR = r"University College London\ISD.ITSD.CO.Technical Specialists - patching"
WINDOWS_WORKBOOK_PATHS = [
    rf"C:\Users\anson\{WORKBOOK_DIR}\{WORKBOOK_NAME}",
    rf"C:\Users\cceadan\{WORKBOOK_DIR}\{WORKBOOK_NAME}",
]


def normalize_workbook_path(value):
    path_text = str(value).strip().strip('"')

    if os.name != "nt":
        m = re.match(r"^([a-zA-Z]):\\(.*)$", path_text)
        if m:
            drive, rest = m.groups()
            rest = rest.replace("\\", "/")
            return Path(f"/mnt/{drive.lower()}/{rest}")

    return Path(path_text)

def first_existing_path(values):
    paths = [normalize_workbook_path(v) for v in values if v]
    return next((p for p in paths if p.exists()), paths[0])

# Fixed workbook path for testing - supports Windows and WSL/Linux
DEFAULT_PATH = first_existing_path([os.getenv("WORKBOOK_PATH"), *WINDOWS_WORKBOOK_PATHS])

# ============================================================
# 2. GENERAL & CELL HELPERS
# ============================================================
def now_str():
    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")

def friendly_now_str():
    dt = datetime.now().astimezone()
    hour = dt.strftime("%I").lstrip("0") or "0"
    return f"{dt.strftime('%A')} {hour}{dt.strftime('%p').lower()} on the {dt.day}/{dt.month}/{dt.year}"

def log(msg):
    print(msg, flush=True)

def clean_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def normalize_header(v):
    return re.sub(r"\s+", " ", clean_text(v).replace("\n", " ")).strip().lower()

def normalize_mac(v):
    return re.sub(r"[^0-9a-fA-F]", "", str(v or "")).lower()

def first_match(pattern, text):
    m = re.search(pattern, text)
    return m.group(0) if m else "Unknown"

def get_lock_owner(path):
    lk = path.parent / f"~${path.name}"
    if not lk.exists():
        return "Unknown (Closed)"
    try:
        return first_match(
            r"[a-zA-Z\s]{3,}",
            lk.read_bytes().decode("latin-1", errors="ignore")
        ).strip() or "a Colleague"
    except Exception:
        return "a Colleague"

def read_only_mount_for(path):
    if os.name == "nt":
        return None

    try:
        path = Path(path).resolve()
        mounts = []
        with open("/proc/mounts", encoding="utf-8") as fh:
            for line in fh:
                parts = line.split()
                if len(parts) >= 4:
                    mount_point = Path(parts[1].replace("\\040", " "))
                    mounts.append((mount_point, parts[3].split(",")))
    except OSError:
        return None

    matches = [
        (mount_point, options)
        for mount_point, options in mounts
        if path == mount_point or mount_point in path.parents
    ]
    if not matches:
        return None

    mount_point, options = max(matches, key=lambda item: len(str(item[0])))
    return mount_point if "ro" in options else None

def assert_workbook_save_ready(path):
    read_only_mount = read_only_mount_for(path)
    if read_only_mount:
        raise RuntimeError(
            f"Workbook is under read-only mount {read_only_mount}. "
            "Restart WSL or remount the drive read-write before running the script."
        )

    try:
        with tempfile.NamedTemporaryFile(prefix=".write-test-", suffix=".tmp", dir=path.parent, delete=True):
            pass
    except PermissionError as exc:
        raise RuntimeError(
            f"Cannot write to workbook folder: {path.parent}. "
            "Close Excel, check OneDrive/SharePoint sync, and make sure the drive is mounted read-write."
        ) from exc
    except OSError as exc:
        raise RuntimeError(f"Workbook folder is not writable: {path.parent} ({exc})") from exc

    try:
        with open(path, "r+b"):
            pass
    except PermissionError as exc:
        raise RuntimeError(
            f"Workbook is not writable: {path}. "
            "Close Excel, wait for OneDrive/SharePoint sync to finish, and check that the file is not marked read-only."
        ) from exc
    except OSError as exc:
        raise RuntimeError(f"Workbook cannot be opened for writing: {path} ({exc})") from exc

def windows_excel_path(path):
    path = Path(path)
    text = str(path)

    if os.name == "nt":
        return text

    match = re.match(r"^/mnt/([a-zA-Z])/(.*)$", text)
    if not match:
        return text

    drive, rest = match.groups()
    rest = rest.replace("/", "\\")
    return f"{drive.upper()}:\\{rest}"

def wait_for_excel_lock_to_clear(path, timeout=30):
    lock_path = path.parent / f"~${path.name}"
    deadline = time.time() + timeout
    while time.time() < deadline:
        if not lock_path.exists():
            return True
        time.sleep(1)
    return False

def get_windows_file_attributes(path):
    if os.name != "nt":
        return None

    try:
        import ctypes
    except ImportError:
        return None

    attrs = ctypes.windll.kernel32.GetFileAttributesW(str(path))
    if attrs == 0xFFFFFFFF:
        return None
    return attrs

def workbook_sync_state(path):
    attrs = get_windows_file_attributes(path)
    if attrs is None:
        return "unknown"

    flags = []
    checks = [
        ("offline", 0x00001000),
        ("recall-on-open", 0x00040000),
        ("recall-on-data-access", 0x00400000),
        ("pinned", 0x00080000),
        ("unpinned", 0x00100000),
    ]
    for name, bit in checks:
        if attrs & bit:
            flags.append(name)

    return ", ".join(flags) if flags else "local"

def should_excel_pre_sync(path):
    if os.getenv("SKIP_EXCEL_PRE_SYNC", "").strip().lower() in {"1", "true", "yes"}:
        return False, "SKIP_EXCEL_PRE_SYNC is set"

    if os.getenv("FORCE_EXCEL_PRE_SYNC", "").strip().lower() in {"1", "true", "yes"}:
        return True, "FORCE_EXCEL_PRE_SYNC is set"

    state = workbook_sync_state(path)
    if any(flag in state for flag in ("offline", "recall-on-open", "recall-on-data-access", "unpinned")):
        return True, f"OneDrive file attributes indicate not fully local: {state}"

    if state == "unknown":
        return False, "sync state unknown; file exists locally, so skipping Excel pre-sync"

    return False, f"OneDrive file attributes look local: {state}"

def open_save_close_in_excel(path):
    lock_path = path.parent / f"~${path.name}"
    if lock_path.exists():
        raise RuntimeError(f"Workbook is already open by {get_lock_owner(path)}.")

    excel_path = windows_excel_path(path)
    log(f"[*] Opening and saving in Excel to refresh SharePoint/OneDrive sync: {path.name}")

    if os.name == "nt":
        try:
            import win32com.client
        except ImportError as exc:
            raise RuntimeError(
                "pywin32 is required for Excel pre-sync on Windows. Install it with: pip install pywin32"
            ) from exc

        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            excel.Visible = False
            wb = excel.Workbooks.Open(excel_path, UpdateLinks=0, ReadOnly=False)
            wb.Save()
        except Exception as exc:
            raise RuntimeError(
                f"Excel could not open/save {path.name}. "
                "This commonly happens when the self-hosted runner is running as a Windows service, "
                "the workbook is already open in another Excel session, or OneDrive has not synced the file "
                f"for the runner user. Excel path: {excel_path}. Details: {exc}"
            ) from exc
        finally:
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
    else:
        ps_script = r"""
$path = $args[0]
$excel = New-Object -ComObject Excel.Application
$excel.DisplayAlerts = $false
$excel.Visible = $false
$workbook = $null
try {
    $workbook = $excel.Workbooks.Open($path, 0, $false)
    $workbook.Save()
} finally {
    if ($workbook -ne $null) {
        $workbook.Close($false)
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }
    $excel.Quit()
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""
        try:
            result = subprocess.run(
                ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script, excel_path],
                capture_output=True,
                text=True,
                timeout=120,
                check=False,
            )
        except FileNotFoundError as exc:
            raise RuntimeError(
                "powershell.exe was not found. Excel pre-sync needs Windows PowerShell when running from WSL."
            ) from exc

        if result.returncode != 0:
            details = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(f"Excel pre-sync failed: {details}")

    if not wait_for_excel_lock_to_clear(path):
        raise RuntimeError(f"Excel lock did not clear after saving {path.name}.")

    log(f"[+] Excel pre-sync complete: {path.name}")

def excel_pre_sync_required():
    return os.getenv("EXCEL_PRE_SYNC_REQUIRED", "").strip().lower() in {"1", "true", "yes"}

def parse_sheet_filter(value):
    return {
        normalize_header(item)
        for item in str(value or "").split(",")
        if item.strip()
    }

def parse_row_filter(value):
    rows = set()
    for item in str(value or "").split(","):
        item = item.strip()
        if not item:
            continue
        if "-" in item:
            start, end = item.split("-", 1)
            if start.strip().isdigit() and end.strip().isdigit():
                lo, hi = int(start), int(end)
                rows.update(range(min(lo, hi), max(lo, hi) + 1))
            continue
        if item.isdigit():
            rows.add(int(item))
    return rows

def confirm_change(safe, ip, port, cur, tgt, row, s_name):
    if not safe:
        return True
    if not os.isatty(0):
        raise RuntimeError("--safe was supplied, but no interactive terminal is available.")
    print(
        f"\n{'='*60}\n"
        f"[ACTION REQUIRED] Sheet {s_name} Row {row}: {ip} {port}\n"
        f"Live VLAN: [{cur}] -> Target VLAN: [{tgt}]"
    )
    return input("Apply change? (y/n): ").strip().lower() == "y"

# ============================================================
# 3. TEAMS ENGINE
# ============================================================
def build_teams_text(status, stats, report):
    lines = [
        f"Run by: {RUN_ACTOR}",
        f"Run at on {friendly_now_str()}",
        f"Source: {RUN_SOURCE}",
        f"File: {DEFAULT_PATH.name}",
        "",
        (
            f"Summary: checked {stats['chk']}, changed {stats['chg']}, "
            f"already correct {stats['ok']}, failed {stats['fail']}, declined {stats['dec']}"
        ),
    ]

    sections = [
        ("Changed", report["changed"]),
        ("Already correct", report["unchanged"]),
        ("Declined", report["declined"]),
        ("Failed", report["failed"]),
    ]

    for title, entries in sections:
        lines.extend(["", f"{title}: {len(entries)}"])
        if not entries:
            lines.append("- None")
            continue

        for e in entries[:25]:
            line = (
                f"- {e['sheet']} row {e['row']} | {e['switch']} | Port {e['port']} | "
                f"{e.get('old_vlan', 'Unknown')} -> {e.get('target_vlan', e.get('live_vlan', 'Unknown'))}"
            )
            if e.get("live_vlan") and e["live_vlan"] != e.get("target_vlan"):
                line += f" | Live: {e['live_vlan']}"
            if e.get("reason"):
                line += f" | {e['reason']}"
            lines.append(line)

        if len(entries) > 25:
            lines.append(f"- ...and {len(entries) - 25} more")

    return "\n".join(lines)

def build_text_card(status, text):
    color = {
        "SUCCESS": "Good",
        "WARNING": "Warning",
        "CRITICAL": "Attention",
    }.get(status, "Accent")

    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": [
            {
                "type": "TextBlock",
                "text": f"Aruba Port Sync: {status}",
                "weight": "Bolder",
                "size": "Medium",
                "color": color,
                "wrap": True,
            },
            {
                "type": "TextBlock",
                "text": text,
                "wrap": True,
            },
        ],
    }

def build_adaptive_card(status, message, details=None):
    return build_text_card(status, message)

def send_teams_notification(status, message, details=None):
    if not TEAMS_WEBHOOK_URL:
        return

    try:
        requests.post(
            TEAMS_WEBHOOK_URL,
            json={"text": message, "adaptive_card": build_text_card(status, message)},
            timeout=10
        ).raise_for_status()
    except Exception as exc:
        log(f"[!] Teams Alert Failed: {exc}")

def record_report(report, bucket, **entry):
    report[bucket].append(entry)

def log_row_failure(row, switch_ip, port, target_vlan, reason, live_vlan="Unknown", old_vlan="Unknown"):
    log(
        f"[FAILED] Row {row} | Switch {switch_ip} | Port {port} | "
        f"Old VLAN {old_vlan} | Target VLAN {target_vlan} | Live VLAN {live_vlan} | Reason: {reason}"
    )

# ============================================================
# 4. EXCEL SCANNING / BLOCK PARSING
# ============================================================
def row_text_map(ws, row):
    return {
        c: clean_text(ws.cell(row=row, column=c).value)
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=row, column=c).value is not None
    }

def is_primary_header_row(ws, row):
    v = {normalize_header(x) for x in row_text_map(ws, row).values()}
    return (
        "vlan" in v
        and "port" in v
        and any(x in v for x in ("switch ip", "switch id"))
        and any(x in v for x in ("data outlet id", "service description", "room name"))
    )

def is_secondary_header_row(ws, row):
    return any(
        x in {normalize_header(v) for v in row_text_map(ws, row).values()}
        for x in ("bldg", "cr", "outlet")
    )

def resolve_block_columns(ws, h_row):
    ex = {}
    norm = defaultdict(list)

    for col, val in row_text_map(ws, h_row).items():
        ex.setdefault(val, col)
        norm[normalize_header(val)].append(col)

    def find_col(keys, is_ex=False, min_c=0):
        if is_ex:
            return next((ex[k] for k in keys if k in ex), None)
        return max(([c for k in keys for c in norm.get(k, []) if c > min_c]), default=None)

    in_vlan = find_col(["VLAN", "Target VLAN", "Requested VLAN"], True) or find_col(
        ["vlan", "target vlan", "requested vlan"]
    )
    sw_ip = find_col(["SWITCH IP", "Switch IP", "Switch ID", "switch ip", "switch id"], True) or find_col(
        ["switch ip", "switch id"]
    )
    in_prt = find_col(["Port", "PORT"], True) or find_col(["port"])

    if not all([in_vlan, sw_ip, in_prt]):
        return {"input_vlan": None, "switch_ip": None, "input_port": None}

    mx_in = max(c for c in (in_vlan, sw_ip, in_prt) if c)

    out = {
        k: find_col(v, min_c=mx_in)
        for k, v in {
            "out_switch": ["switch", "configured switch", "live switch"],
            "out_port": ["port", "configured port", "live port"],
            "out_vlan": ["vlan", "configured vlan", "live vlan", "done vlan"],
            "out_mac": ["mac", "mac address"],
            "out_ip": ["ip", "ip address"],
            "out_time": ["time", "last checked", "checked at"],
        }.items()
    }

    out["out_notes"] = find_col(["notes", "status", "Status"], True) or find_col(
        ["status", "notes"], min_c=mx_in
    )

    if out["out_notes"] and any(c and c > out["out_notes"] for c in out.values() if c != out["out_notes"]):
        out["out_notes"] = None

    return {"input_vlan": in_vlan, "switch_ip": sw_ip, "input_port": in_prt, **out}

def collect_sheet_blocks(ws):
    blocks = []
    h_rows = [r for r in range(1, ws.max_row + 1) if is_primary_header_row(ws, r)]

    for idx, hr in enumerate(h_rows):
        nxt = h_rows[idx + 1] if idx + 1 < len(h_rows) else ws.max_row + 1
        cols = resolve_block_columns(ws, hr)

        if all(cols[x] for x in ["input_vlan", "switch_ip", "input_port"]):
            title = "Unknown"
            for r in range(hr - 1, max(hr - 3, 0), -1):
                v = clean_text(ws.cell(row=r, column=1).value)
                if v and normalize_header(v) not in {"bldg", "data outlet id"}:
                    title = v
                    break

            blocks.append({
                "sheet_name": ws.title,
                "section_name": title if title != "Unknown" else ws.title,
                "header_row": hr,
                "data_start": hr + 2 if is_secondary_header_row(ws, hr + 1) else hr + 1,
                "data_end": nxt - 1,
                "columns": cols
            })

    return blocks

def write_result_columns(ws, row, cols, **kwargs):
    mapping = {
        "out_switch": kwargs.get("switch_ip"),
        "out_port": kwargs.get("port"),
        "out_vlan": kwargs.get("vlan"),
        "out_mac": kwargs.get("mac"),
        "out_ip": kwargs.get("ip"),
        "out_time": kwargs.get("checked_at"),
        "out_notes": kwargs.get("notes"),
    }
    for k, v in mapping.items():
        if cols.get(k):
            ws.cell(row=row, column=cols[k], value=v)

def is_yellow_fill(fill):
    if not fill or not fill.fill_type:
        return False

    for color in (fill.fgColor, fill.start_color):
        if not color:
            continue
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb.upper()[-6:]
            try:
                red = int(rgb[0:2], 16)
                green = int(rgb[2:4], 16)
                blue = int(rgb[4:6], 16)
            except ValueError:
                continue
            if red >= 220 and green >= 180 and blue <= 140:
                return True
        if color.type == "indexed" and color.indexed in {6, 27, 36, 44}:
            return True
        if color.type == "theme" and fill.fill_type == "solid":
            return True

    return False

def is_highlighted_row(ws, row):
    if is_yellow_fill(ws.row_dimensions[row].fill):
        return True
    return any(is_yellow_fill(ws.cell(row=row, column=c).fill) for c in range(1, ws.max_column + 1))

def fill_debug_summary(ws, row):
    parts = []
    row_fill = ws.row_dimensions[row].fill
    if row_fill and row_fill.fill_type:
        parts.append(
            f"row-fill type={row_fill.fill_type} "
            f"fg={row_fill.fgColor.type}:{row_fill.fgColor.rgb or row_fill.fgColor.indexed or row_fill.fgColor.theme} "
            f"start={row_fill.start_color.type}:{row_fill.start_color.rgb or row_fill.start_color.indexed or row_fill.start_color.theme}"
        )

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        fill = cell.fill
        if not fill or not fill.fill_type:
            continue
        parts.append(
            f"{cell.coordinate} type={fill.fill_type} "
            f"fg={fill.fgColor.type}:{fill.fgColor.rgb or fill.fgColor.indexed or fill.fgColor.theme} "
            f"start={fill.start_color.type}:{fill.start_color.rgb or fill.start_color.indexed or fill.start_color.theme}"
        )

    return "; ".join(parts[:20]) if parts else "no explicit row/cell fills found by openpyxl"

def clear_yellow_highlight(ws, row):
    if is_yellow_fill(ws.row_dimensions[row].fill):
        ws.row_dimensions[row].fill = PatternFill(fill_type=None)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        if is_yellow_fill(cell.fill):
            cell.fill = PatternFill(fill_type=None)

# ============================================================
# 5. SWITCH INFRASTRUCTURE LOGIC
# ============================================================
def run_cmd_safe(conn, cmd, t=30):
    try:
        return conn.send_command(cmd, read_timeout=t)
    except Exception:
        return ""

def get_port_live_details(conn, port):
    mac = "Unknown"

    for cmd in [f"show mac-address-table interface {port}", f"show mac-address-table int {port}"]:
        mac = first_match(
            r"(?:[0-9a-fA-F]{2}[:.-]){5}[0-9a-fA-F]{2}|(?:[0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4}",
            run_cmd_safe(conn, cmd)
        )
        if mac != "Unknown":
            break

    if mac == "Unknown":
        mac = next(
            (
                first_match(
                    r"(?:[0-9a-fA-F]{2}[:.-]){5}[0-9a-fA-F]{2}|(?:[0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4}",
                    l
                )
                for l in run_cmd_safe(conn, "show mac-address-table", 60).splitlines()
                if port in l
            ),
            "Unknown"
        )

    ip, m_norm = "Unknown", normalize_mac(mac)
    if m_norm:
        for cmd in ("show arp", "show arp all-vrfs"):
            ip = next(
                (
                    first_match(r"\b(?:\d{1,3}\.){3}\d{1,3}\b", l)
                    for l in run_cmd_safe(conn, cmd, 60).splitlines()
                    if m_norm in normalize_mac(l)
                ),
                "Unknown"
            )
            if ip != "Unknown":
                break

    return {"mac": mac, "ip": ip}

def parse_switch_port(port):
    m = re.match(r"^(\d+/\d+/)(\d+)$", clean_text(port))
    if not m:
        return None
    prefix, number = m.groups()
    return prefix, int(number)

def build_vlan_change_groups(items):
    groups = []
    sortable = defaultdict(list)

    for idx, item in enumerate(items):
        parsed = parse_switch_port(item["port"])
        if parsed:
            prefix, number = parsed
            sortable[(item["target_vlan"], prefix)].append((number, idx, item))
        else:
            groups.append({
                "order": idx,
                "interface": item["port"],
                "target_vlan": item["target_vlan"],
                "items": [item],
            })

    for (vlan, prefix), ports in sortable.items():
        ports.sort(key=lambda x: x[0])
        start_num, start_idx, first_item = ports[0]
        prev_num, prev_idx = start_num, start_idx
        current_items = [first_item]

        for number, idx, item in ports[1:]:
            if number == prev_num + 1:
                current_items.append(item)
                prev_num, prev_idx = number, idx
                continue

            interface = (
                f"{prefix}{start_num}"
                if start_num == prev_num
                else f"{prefix}{start_num}-{prefix}{prev_num}"
            )
            groups.append({
                "order": start_idx,
                "interface": interface,
                "target_vlan": vlan,
                "items": current_items,
            })
            start_num, start_idx = number, idx
            prev_num, prev_idx = number, idx
            current_items = [item]

        interface = (
            f"{prefix}{start_num}"
            if start_num == prev_num
            else f"{prefix}{start_num}-{prefix}{prev_num}"
        )
        groups.append({
            "order": start_idx,
            "interface": interface,
            "target_vlan": vlan,
            "items": current_items,
        })

    return sorted(groups, key=lambda x: x["order"])

CLI_ERROR_PATTERNS = (
    "Invalid input",
    "Incomplete command",
    "Ambiguous command",
    "Unknown command",
    "Command failed",
    "Error:",
)

def compact_cli_output(text):
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    return " | ".join(lines)[-300:]

def find_cli_error(output):
    for line in str(output or "").splitlines():
        if any(pattern.lower() in line.lower() for pattern in CLI_ERROR_PATTERNS):
            return line.strip()
    return ""

def apply_vlan_change(conn, interface, vlan):
    for cmd in ["configure terminal", f"interface {interface}", f"vlan access {vlan}", "end"]:
        output = conn.send_command_timing(cmd)
        error = find_cli_error(output)
        if error:
            raise RuntimeError(f"Command `{cmd}` failed: {error} | Output: {compact_cli_output(output)}")

# ============================================================
# 6. PIPELINE CONTROLLER
# ============================================================
def main():
    global DEFAULT_PATH

    parser = argparse.ArgumentParser()
    parser.add_argument("--safe", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--workbook",
        default=None,
        help="Workbook path. Accepts Windows paths under WSL, for example C:\\Users\\anson\\...\\FC-MSA-CI.xlsx",
    )
    parser.add_argument(
        "--sheets",
        default=os.getenv("SHEETS_TO_PROCESS", "Foster Court"),
        help="Comma-separated sheet names to process. Default: Foster Court. Use '*' with --sheets or SHEETS_TO_PROCESS to process all sheets.",
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Process every worksheet instead of only the configured sheet filter.",
    )
    parser.add_argument(
        "--rows",
        default=os.getenv("ROWS_TO_PROCESS", ""),
        help="Optional comma-separated row numbers or ranges to process even if yellow fill is not detected, for example 30,31 or 30-35.",
    )
    args = parser.parse_args()
    if args.workbook:
        DEFAULT_PATH = normalize_workbook_path(args.workbook)

    log(f"[*] Script starting\n[*] Safe mode: {args.safe}\n[*] Dry run: {args.dry_run}\n[*] Path: {DEFAULT_PATH}")
    log(f"[*] Python executable: {sys.executable}")
    log(f"[*] Platform: {platform.platform()}")
    log(f"[*] os.name: {os.name}")
    log(f"[*] Workbook exists check: {DEFAULT_PATH.exists()}")

    if not DEFAULT_PATH.exists():
        log(f"[!] Workbook not found: {DEFAULT_PATH}")
        sys.exit(1)

    if not USERNAME:
        log("[!] Missing environment variable: username")
        sys.exit(1)

    if not PASSWORD:
        log("[!] Missing environment variable: passwordAD")
        sys.exit(1)

    pre_sync_warning = ""
    needs_pre_sync, pre_sync_reason = should_excel_pre_sync(DEFAULT_PATH)
    log(f"[*] Excel pre-sync check: {pre_sync_reason}")
    if needs_pre_sync:
        try:
            open_save_close_in_excel(DEFAULT_PATH)
        except RuntimeError as exc:
            if excel_pre_sync_required():
                log(f"[!] ABORTED: {exc}")
                sys.exit(1)
            pre_sync_warning = str(exc)
            log(f"[!] Excel pre-sync warning: {pre_sync_warning}")
            log("[*] Continuing because EXCEL_PRE_SYNC_REQUIRED is not set.")

    lk = DEFAULT_PATH.parent / f"~${DEFAULT_PATH.name}"
    if lk.exists():
        log(f"[!] ABORTED: Open by {get_lock_owner(DEFAULT_PATH)}.")
        sys.exit(1)

    try:
        assert_workbook_save_ready(DEFAULT_PATH)
    except RuntimeError as exc:
        log(f"[!] ABORTED: {exc}")
        sys.exit(1)

    start_mtime = os.path.getmtime(DEFAULT_PATH)

    try:
        wb = load_workbook(DEFAULT_PATH, data_only=False)
        log(f"[*] Workbook loaded successfully: {DEFAULT_PATH.name}")
        log(f"[*] Sheets found: {wb.sheetnames}")
    except Exception as exc:
        log(f"[!] Load error: {exc}")
        sys.exit(1)

    sheet_filter = parse_sheet_filter(args.sheets)
    process_all_sheets = args.all_sheets or "*" in sheet_filter or "all" in sheet_filter
    if process_all_sheets:
        worksheets_to_process = list(wb.worksheets)
        log("[*] Sheet filter: all worksheets")
    else:
        worksheets_to_process = [
            ws for ws in wb.worksheets if normalize_header(ws.title) in sheet_filter
        ]
        log(f"[*] Sheet filter: {', '.join(item.strip() for item in args.sheets.split(',') if item.strip())}")

        if not worksheets_to_process:
            log(f"[!] No worksheets matched sheet filter: {args.sheets}")
            log(f"[*] Available worksheets: {wb.sheetnames}")
            sys.exit(1)

        skipped_sheet_names = [ws.title for ws in wb.worksheets if ws not in worksheets_to_process]
        if skipped_sheet_names:
            log(f"[*] Skipping worksheets by filter: {skipped_sheet_names}")

    all_blocks = [(ws, b) for ws in worksheets_to_process for b in collect_sheet_blocks(ws)]
    log(f"[*] Total blocks found: {len(all_blocks)}")

    if not all_blocks:
        log("[*] No blocks.")
        sys.exit(0)

    stats = {"chk": 0, "ok": 0, "chg": 0, "fail": 0, "dec": 0}
    wb_touch = False
    report = {"changed": [], "unchanged": [], "declined": [], "failed": []}

    if args.dry_run:
        log("!!!!!!!!!!!!!!!!!!!! DRY RUN ACTIVE !!!!!!!!!!!!!!!!!!!!")

    forced_rows = parse_row_filter(args.rows)
    if forced_rows:
        log(f"[*] Forced row filter active: {sorted(forced_rows)}")

    blocks_by_sheet = defaultdict(list)
    for ws, b in all_blocks:
        blocks_by_sheet[ws.title].append(b)

    for ws in worksheets_to_process:
        covered_rows = set()
        for b in blocks_by_sheet.get(ws.title, []):
            covered_rows.update(range(b["data_start"], b["data_end"] + 1))

        outside_rows = [
            r
            for r in range(1, ws.max_row + 1)
            if (is_highlighted_row(ws, r) or r in forced_rows) and r not in covered_rows
        ]
        if not outside_rows:
            continue

        log(
            f"[!] Highlighted/forced row(s) on sheet '{ws.title}' are outside detected data blocks "
            f"and will not be processed: {outside_rows}"
        )
        for r in outside_rows:
            reason = "Highlighted/forced row is outside detected data blocks"
            log(f"[*] Ignored row {r} on sheet '{ws.title}': {reason}")
            if r in forced_rows:
                log(f"[*] Fill debug for {ws.title} row {r}: {fill_debug_summary(ws, r)}")

    for ws, b in all_blocks:
        log(
            f"[*] Processing sheet '{b['sheet_name']}' "
            f"section '{b['section_name']}' "
            f"rows {b['data_start']} to {b['data_end']}"
        )

        cols = b["columns"]
        rows_by_sw = defaultdict(list)
        highlighted_rows = 0
        queued_highlights = []
        skipped_highlights = []

        for r in range(b["data_start"], b["data_end"] + 1):
            highlighted = is_highlighted_row(ws, r)
            forced = r in forced_rows
            if not highlighted and not forced:
                continue

            if forced:
                log(
                    f"[*] Forced row {r} on sheet '{ws.title}' selected. "
                    f"Yellow detected: {highlighted}. Fill debug: {fill_debug_summary(ws, r)}"
                )

            sw, pt, tg = [
                clean_text(ws.cell(row=r, column=cols[k]).value)
                for k in ["switch_ip", "input_port", "input_vlan"]
            ]
            if sw and pt and tg:
                highlighted_rows += 1
                source = "yellow" if highlighted else "forced"
                queued_highlights.append(f"{r} [{source}] ({sw} {pt} -> VLAN {tg})")
                rows_by_sw[sw].append({
                    "row_idx": r,
                    "port": pt,
                    "target_vlan": tg,
                    "cell": ws.cell(row=r, column=cols["input_vlan"])
                })
            else:
                missing = [
                    name
                    for name, value in [("switch", sw), ("port", pt), ("target VLAN", tg)]
                    if not value
                ]
                reason = f"Highlighted row skipped because missing {', '.join(missing)}"
                skipped_highlights.append(f"{r} ({reason})")
                stats["fail"] += 1
                log_row_failure(
                    r,
                    sw or "Unknown",
                    pt or "Unknown",
                    tg or "Unknown",
                    reason,
                )
                record_report(
                    report,
                    "failed",
                    sheet=b["sheet_name"],
                    row=r,
                    switch=sw or "Unknown",
                    port=pt or "Unknown",
                    old_vlan="Unknown",
                    target_vlan=tg or "Unknown",
                    live_vlan="Unknown",
                    reason=reason,
                )
                if not args.dry_run:
                    write_result_columns(
                        ws,
                        r,
                        cols,
                        switch_ip=sw or "Unknown",
                        port=pt or "Unknown",
                        vlan="Unknown",
                        mac="N/A",
                        ip="N/A",
                        checked_at=now_str(),
                        notes=reason[:255],
                    )
                    wb_touch = True

        log(f"[*] Highlighted rows queued in this block: {highlighted_rows}")
        if queued_highlights:
            log(f"[*] Queued highlighted row detail: {queued_highlights}")
        if skipped_highlights:
            log(f"[!] Highlighted rows skipped due to missing switch/port/VLAN: {skipped_highlights}")
        log(f"[*] Unique switches in this block: {len(rows_by_sw)}")

        for sw_ip, entries in rows_by_sw.items():
            log(f"[*] Connecting to switch: {sw_ip} ({len(entries)} row(s))")

            try:
                with ConnectHandler(
                    device_type="aruba_aoscx",
                    host=sw_ip,
                    username=USERNAME,
                    password=PASSWORD,
                    conn_timeout=20,
                    fast_cli=False
                ) as conn:
                    log(f"[+] Connected to {sw_ip}")

                    try:
                        conn.send_command_timing("no page")
                        conn.send_command_timing("aruba-central support-mode")
                    except Exception:
                        pass

                    cur_v_map = {}
                    for l in conn.send_command("show int br", read_timeout=60).splitlines():
                        m = re.match(r"^\s*(\d+/\d+/\d+)\s+(\S+)", l.rstrip())
                        if m:
                            cur_v_map[m.group(1).strip()] = m.group(2).strip()

                    approved_changes = []
                    pending = []

                    for e in entries:
                        r = e["row_idx"]
                        pt = e["port"]
                        tg = e["target_vlan"]
                        cur_v = cur_v_map.get(pt, "Unknown")
                        stats["chk"] += 1

                        log(f"[*] Row {r} | Port {pt} | Current VLAN {cur_v} | Target VLAN {tg}")

                        if cur_v == tg:
                            stats["ok"] += 1
                            log(f"[OK] Row {r} already on correct VLAN")
                            record_report(
                                report,
                                "unchanged",
                                sheet=b["sheet_name"],
                                row=r,
                                switch=sw_ip,
                                port=pt,
                                old_vlan=cur_v,
                                target_vlan=tg,
                                live_vlan=cur_v,
                                reason="No change needed",
                            )
                            if not args.dry_run:
                                for k, v in {
                                    "out_switch": sw_ip,
                                    "out_port": pt,
                                    "out_vlan": cur_v,
                                    "out_time": now_str(),
                                    "out_notes": "No change needed",
                                }.items():
                                    if cols.get(k):
                                        ws.cell(row=r, column=cols[k], value=v)
                                clear_yellow_highlight(ws, r)
                                wb_touch = True
                            continue

                        if args.dry_run:
                            log(f"[DRY-RUN] Row {r} | {sw_ip} | Port {pt} | {cur_v} -> {tg}")
                            continue

                        if not confirm_change(args.safe, sw_ip, pt, cur_v, tg, r, b["sheet_name"]):
                            dt = get_port_live_details(conn, e["port"])
                            stats["dec"] += 1
                            reason = "Change declined in safe mode"
                            log(f"[SKIPPED] Row {r} | Switch {sw_ip} | Port {pt} | Reason: {reason}")
                            record_report(
                                report,
                                "declined",
                                sheet=b["sheet_name"],
                                row=r,
                                switch=sw_ip,
                                port=pt,
                                old_vlan=cur_v,
                                target_vlan=tg,
                                live_vlan=cur_v,
                                reason=reason,
                            )
                            write_result_columns(
                                ws,
                                r,
                                cols,
                                switch_ip=sw_ip,
                                port=pt,
                                vlan=cur_v,
                                mac=dt["mac"],
                                ip=dt["ip"],
                                checked_at=now_str(),
                                notes=f"Current VLAN: {cur_v}"
                            )
                            wb_touch = True
                            continue

                        approved_changes.append({
                            "row_idx": r,
                            "port": pt,
                            "target_vlan": tg,
                            "old_vlan": cur_v,
                            "cell": e["cell"]
                        })

                    for group in build_vlan_change_groups(approved_changes):
                        rows = ", ".join(str(item["row_idx"]) for item in group["items"])

                        try:
                            log(
                                f"[*] Applying change on {sw_ip} {group['interface']}: "
                                f"VLAN {group['target_vlan']} (row(s): {rows})"
                            )
                            apply_vlan_change(conn, group["interface"], group["target_vlan"])
                            pending.extend(group["items"])
                        except Exception as exc:
                            reason = str(exc)[:250]
                            for item in group["items"]:
                                stats["fail"] += 1
                                log_row_failure(
                                    item["row_idx"],
                                    sw_ip,
                                    item["port"],
                                    item["target_vlan"],
                                    reason,
                                    live_vlan=item["old_vlan"],
                                    old_vlan=item["old_vlan"],
                                )
                                record_report(
                                    report,
                                    "failed",
                                    sheet=b["sheet_name"],
                                    row=item["row_idx"],
                                    switch=sw_ip,
                                    port=item["port"],
                                    old_vlan=item["old_vlan"],
                                    target_vlan=item["target_vlan"],
                                    live_vlan=item["old_vlan"],
                                    reason=f"Apply failed: {reason}",
                                )
                                write_result_columns(
                                    ws,
                                    item["row_idx"],
                                    cols,
                                    switch_ip=sw_ip,
                                    port=item["port"],
                                    vlan=item["old_vlan"],
                                    mac="Unknown",
                                    ip="Unknown",
                                    checked_at=now_str(),
                                    notes=f"Error: {str(exc)[:50]}"
                                )
                                wb_touch = True

                    if pending:
                        log(f"[*] Verifying {len(pending)} changed row(s) on {sw_ip}")

                        post_v_map = {}
                        for l in conn.send_command("show int br", read_timeout=60).splitlines():
                            m = re.match(r"^\s*(\d+/\d+/\d+)\s+(\S+)", l.rstrip())
                            if m:
                                post_v_map[m.group(1).strip()] = m.group(2).strip()

                        for item in pending:
                            v_fin = post_v_map.get(item["port"], "Unknown")
                            dt = get_port_live_details(conn, item["port"])

                            if v_fin == item["target_vlan"]:
                                stats["chg"] += 1
                                log(f"[DONE] Row {item['row_idx']} verified successfully")
                                write_result_columns(
                                    ws,
                                    item["row_idx"],
                                    cols,
                                    switch_ip=sw_ip,
                                    port=item["port"],
                                    vlan=v_fin,
                                    mac=dt["mac"],
                                    ip=dt["ip"],
                                    checked_at=now_str()
                                )
                                clear_yellow_highlight(ws, item["row_idx"])
                                record_report(report, "changed", **{
                                    "sheet": b["sheet_name"],
                                    "row": item["row_idx"],
                                    "switch": sw_ip,
                                    "port": item["port"],
                                    "target_vlan": item["target_vlan"],
                                    "old_vlan": item["old_vlan"],
                                    "live_vlan": v_fin,
                                    "reason": "Verified after apply",
                                })
                            else:
                                stats["fail"] += 1
                                if v_fin == "Unknown":
                                    failure_reason = f"port {item['port']} was not found in `show int br` after apply"
                                elif v_fin == item["old_vlan"]:
                                    failure_reason = "switch still reports the old VLAN after apply"
                                else:
                                    failure_reason = "switch reports a different VLAN than requested"

                                log_row_failure(
                                    item["row_idx"],
                                    sw_ip,
                                    item["port"],
                                    item["target_vlan"],
                                    failure_reason,
                                    live_vlan=v_fin,
                                    old_vlan=item["old_vlan"],
                                )
                                record_report(
                                    report,
                                    "failed",
                                    sheet=b["sheet_name"],
                                    row=item["row_idx"],
                                    switch=sw_ip,
                                    port=item["port"],
                                    old_vlan=item["old_vlan"],
                                    target_vlan=item["target_vlan"],
                                    live_vlan=v_fin,
                                    reason=f"Verify failed: {failure_reason}",
                                )
                                write_result_columns(
                                    ws,
                                    item["row_idx"],
                                    cols,
                                    switch_ip=sw_ip,
                                    port=item["port"],
                                    vlan=v_fin,
                                    mac=dt["mac"],
                                    ip=dt["ip"],
                                    checked_at=now_str(),
                                    notes=(
                                        f"Failed to change: expected VLAN {item['target_vlan']}, "
                                        f"live VLAN {v_fin}. {failure_reason}"
                                    )[:255]
                                )

                            wb_touch = True

            except Exception as exc:
                reason = str(exc)[:250]
                log(f"[!] Switch connection/processing failure on {sw_ip}: {reason}")
                stats["fail"] += len(entries)
                if not args.dry_run:
                    for e in entries:
                        log_row_failure(
                            e["row_idx"],
                            sw_ip,
                            e["port"],
                            e["target_vlan"],
                            f"Switch connection/processing failed: {reason}",
                        )
                        record_report(
                            report,
                            "failed",
                            sheet=b["sheet_name"],
                            row=e["row_idx"],
                            switch=sw_ip,
                            port=e["port"],
                            old_vlan="Unknown",
                            target_vlan=e["target_vlan"],
                            live_vlan="Unknown",
                            reason=f"Switch connection/processing failed: {reason}",
                        )
                        write_result_columns(
                            ws,
                            e["row_idx"],
                            cols,
                            switch_ip=sw_ip,
                            port=e["port"],
                            vlan="Unknown",
                            mac="Error",
                            ip="Error",
                            checked_at=now_str(),
                            notes="Switch connection failed"
                        )
                    wb_touch = True

    if args.dry_run:
        log("[*] Dry run complete")
        sys.exit(0)

    if wb_touch:
        if os.path.getmtime(DEFAULT_PATH) != start_mtime:
            send_teams_notification("CRITICAL", f"Conflict! Opened by {get_lock_owner(DEFAULT_PATH)}")
            sys.exit(1)
        try:
            wb.save(DEFAULT_PATH)
            log(f"[+] Workbook saved successfully: {DEFAULT_PATH}")
        except Exception as e:
            log(f"[!] Save Error: {e}")
            sys.exit(1)

    st = "CRITICAL" if stats["fail"] > 0 else "SUCCESS" if stats["chg"] > 0 else "INFO"
    log(
        f"[*] Final summary | Checked: {stats['chk']} | Already OK: {stats['ok']} | "
        f"Changed: {stats['chg']} | Failed: {stats['fail']} | Declined: {stats['dec']}"
    )

    for bucket, label in [
        ("changed", "Changed"),
        ("unchanged", "Already correct"),
        ("declined", "Declined"),
        ("failed", "Failed"),
    ]:
        for e in report[bucket]:
            log(
                f"[REPORT] {label} | {e['sheet']} row {e['row']} | Switch {e['switch']} | "
                f"Port {e['port']} | Old {e.get('old_vlan', 'Unknown')} | "
                f"Target {e.get('target_vlan', 'Unknown')} | Live {e.get('live_vlan', 'Unknown')} | "
                f"{e.get('reason', '')}"
            )

    final_message = build_teams_text(st, stats, report)
    if pre_sync_warning:
        final_message = (
            f"{final_message}\n\n"
            f"Excel pre-sync warning:\n{pre_sync_warning}\n"
            "The script continued after the normal workbook access checks."
        )

    send_teams_notification(
        st,
        final_message,
    )

if __name__ == "__main__":
    main()
