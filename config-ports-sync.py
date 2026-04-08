import os
import re
import sys
import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from netmiko import ConnectHandler
from openpyxl import load_workbook

# ============================================================
# 1. SETUP & CONFIGURATION
# ============================================================
load_dotenv()

USERNAME = os.getenv("username")
PASSWORD = os.getenv("passwordAD")
TEAMS_WEBHOOK_URL = os.getenv("TEAMS_WEBHOOK_URL", "")

RUN_ACTOR = (
    os.getenv("GITHUB_ACTOR")
    or os.getenv("USER")
    or os.getenv("USERNAME")
    or "unknown"
)

RUN_SOURCE = (
    "GitHub Actions"
    if os.getenv("GITHUB_ACTIONS", "").lower() == "true"
    else "Manual"
)

# --- PATH AUTO-DISCOVERY ---
ONEDRIVE_PATH = Path(
    "/mnt/c/Users/cceadan/OneDrive - University College London/Estates IT - Project Documentation - Patching Schedule/90TCR - Daniel Test.xlsx"
    #"/mnt/c/Users/cceadan/OneDrive - University College London/Estates IT - Project Documentation - Patching Schedule/90TCR - Level 3B - Patching Schedule.xlsx"
)

DOWNLOADS_PATH = Path(
    "/mnt/c/Users/cceadan/Downloads/90 TCR - Daniel-Test.xlsx"
    "/mnt/c/Users/cceadan/Downloads/90 TCR - Daniel-Test.xlsx"
)

if ONEDRIVE_PATH.exists():
    DEFAULT_PATH = ONEDRIVE_PATH
elif DOWNLOADS_PATH.exists():
    DEFAULT_PATH = DOWNLOADS_PATH
else:
    DEFAULT_PATH = ONEDRIVE_PATH


# ============================================================
# 2. GENERAL HELPERS
# ============================================================
def now_str():
    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def log(message):
    print(message, flush=True)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value):
    text = str(value or "").replace("\n", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_mac(value):
    return re.sub(r"[^0-9a-fA-F]", "", str(value or "")).lower()


def first_mac_in_text(text):
    patterns = [
        r"(?:[0-9a-fA-F]{2}[:.-]){5}[0-9a-fA-F]{2}",
        r"(?:[0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4}",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0)
    return "Unknown"


def first_ip_in_text(text):
    match = re.search(r"\b(?:\d{1,3}\.){3}\d{1,3}\b", text)
    return match.group(0) if match else "Unknown"


def get_lock_owner(path):
    lock_path = path.parent / f"~${path.name}"
    if not lock_path.exists():
        return "Unknown (Closed)"

    try:
        with open(lock_path, "rb") as file_handle:
            content = file_handle.read().decode("latin-1", errors="ignore")
            match = re.search(r"[a-zA-Z\s]{3,}", content)
            return match.group(0).strip() if match else "a Colleague"
    except Exception:
        return "a Colleague"


def confirm_change(safe_mode, switch_ip, port, current_vlan, target_vlan, row_idx):
    if not safe_mode:
        return True

    if not os.isatty(0):
        raise RuntimeError("--safe was supplied, but no interactive terminal is available.")

    print("\n" + "=" * 60)
    print(f"[ACTION REQUIRED] Row {row_idx}: {switch_ip} {port}")
    print(f"Live VLAN: [{current_vlan}] -> Target VLAN: [{target_vlan}]")
    reply = input("Apply change? (y/n): ").strip().lower()
    return reply == "y"


# ============================================================
# 3. TEAMS
# ============================================================
def send_teams_notification(status, message, details=None):
    if not TEAMS_WEBHOOK_URL:
        return

    color = {
        "SUCCESS": "28A745",
        "WARNING": "FFC107",
        "CRITICAL": "DC3545",
    }.get(status, "0078D7")

    facts = [
        {"name": "Run By", "value": RUN_ACTOR},
        {"name": "Run At", "value": now_str()},
        {"name": "Source", "value": RUN_SOURCE},
        {"name": "File", "value": DEFAULT_PATH.name},
    ]

    if details:
        for entry in details:
            facts.append(
                {
                    "name": f"Switch {entry['ip']}",
                    "value": (
                        f"Port {entry['port']} -> VLAN {entry['target_vlan']}\n"
                        f"Previous Live VLAN: {entry['old_vlan']}\n"
                        f"By: {entry['changed_by']}\n"
                        f"At: {entry['changed_at']}\n"
                        f"Source: {entry['source']}"
                    ),
                }
            )

    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": color,
        "summary": "Aruba Config Update",
        "sections": [
            {
                "activityTitle": f"**Aruba Configurator: {status}**",
                "activitySubtitle": f"File: {DEFAULT_PATH.name}",
                "text": message,
                "facts": facts,
            }
        ],
    }

    try:
        response = requests.post(TEAMS_WEBHOOK_URL, json=payload, timeout=10)
        response.raise_for_status()
    except Exception as exc:
        log(f"[!] Teams Alert Failed: {exc}")


# ============================================================
# 4. EXCEL HEADER MAPPING
# ============================================================
def build_header_map(ws):
    """
    Expected useful columns:
      G = VLAN
      H = SWITCH IP
      I = PORT
      L = mac
      M = ip
      N = Last Checked
      O = notes

    We scan rows 2 and 3 and keep the rightmost match, so the lowercase
    notes column on O wins over the earlier Notes column.
    """
    aliases = {
        "target_vlan": {"vlan"},
        "switch_ip": {"switch ip"},
        "port": {"port"},
        "mac": {"mac"},
        "ip": {"ip"},
        "last_checked": {"last checked"},
        "notes": {"notes"},
        "room_name": {"room name", "room"},
        "description": {"description"},
        "outlet": {"outlet"},
        "desk_no": {"desk no"},
        "bldg": {"bldg"},
        "cr": {"cr"},
    }

    header_map = {}

    for col in range(1, ws.max_column + 1):
        for row in (2, 3):
            header = normalize_header(ws.cell(row=row, column=col).value)
            if not header:
                continue

            for key, names in aliases.items():
                if header in names:
                    header_map[key] = col

    return header_map


def write_readonly_columns(ws, row_idx, header_map, mac="", ip="", last_checked="", notes=""):
    if "mac" in header_map:
        ws.cell(row=row_idx, column=header_map["mac"], value=mac)
    if "ip" in header_map:
        ws.cell(row=row_idx, column=header_map["ip"], value=ip)
    if "last_checked" in header_map:
        ws.cell(row=row_idx, column=header_map["last_checked"], value=last_checked)
    if "notes" in header_map:
        ws.cell(row=row_idx, column=header_map["notes"], value=notes)


# ============================================================
# 5. SWITCH HELPERS (ARUBA CX)
# ============================================================
def connect_to_switch(switch_ip):
    device = {
        "device_type": "aruba_aoscx",
        "host": switch_ip,
        "username": USERNAME,
        "password": PASSWORD,
        "conn_timeout": 20,
        "fast_cli": False,
    }
    return ConnectHandler(**device)


def prepare_switch_session(net_connect):
    # Best effort only
    try:
        net_connect.send_command_timing("no page")
    except Exception:
        pass

    try:
        net_connect.send_command_timing("aruba-central support-mode")
    except Exception:
        pass


def parse_show_int_br(output):
    """
    Parse lines like:
    1/1/12         915     access ...
    Returns: {"1/1/12": "915", ...}
    """
    port_vlan_map = {}

    for line in output.splitlines():
        line = line.rstrip()
        match = re.match(r"^\s*(\d+/\d+/\d+)\s+(\S+)", line)
        if match:
            port = match.group(1).strip()
            vlan = match.group(2).strip()
            port_vlan_map[port] = vlan

    return port_vlan_map


def get_live_vlan_map(net_connect):
    output = net_connect.send_command("show int br", read_timeout=60)
    return parse_show_int_br(output)


def get_port_mac(net_connect, port):
    commands = [
        f"show mac-address-table interface {port}",
        f"show mac-address-table int {port}",
    ]

    for cmd in commands:
        try:
            output = net_connect.send_command(cmd, read_timeout=30)
            mac = first_mac_in_text(output)
            if mac != "Unknown":
                return mac
        except Exception:
            continue

    # fallback: full table
    try:
        output = net_connect.send_command("show mac-address-table", read_timeout=60)
        for line in output.splitlines():
            if port in line:
                mac = first_mac_in_text(line)
                if mac != "Unknown":
                    return mac
    except Exception:
        pass

    return "Unknown"


def get_ip_for_mac(net_connect, mac):
    if mac == "Unknown":
        return "Unknown"

    mac_norm = normalize_mac(mac)
    if not mac_norm:
        return "Unknown"

    for cmd in ("show arp", "show arp all-vrfs"):
        try:
            output = net_connect.send_command(cmd, read_timeout=60)
        except Exception:
            continue

        for line in output.splitlines():
            if mac_norm and mac_norm in normalize_mac(line):
                ip = first_ip_in_text(line)
                if ip != "Unknown":
                    return ip

    return "Unknown"


def get_port_live_details(net_connect, port):
    mac = get_port_mac(net_connect, port)
    ip = get_ip_for_mac(net_connect, mac)
    return {"mac": mac, "ip": ip}


def apply_vlan_change(net_connect, port, target_vlan):
    """
    Uses timing-based sends to avoid prompt-detection issues.
    """
    output = ""
    output += net_connect.send_command_timing("configure terminal")
    output += net_connect.send_command_timing(f"interface {port}")
    output += net_connect.send_command_timing(f"vlan access {target_vlan}")
    output += net_connect.send_command_timing("end")
    return output


# ============================================================
# 6. MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--safe", action="store_true", help="Confirm each VLAN change manually")
    parser.add_argument("--dry-run", action="store_true", help="Show what would change without applying it")
    args = parser.parse_args()

    log("[*] Script starting")
    log(f"[*] Safe mode: {args.safe}")
    log(f"[*] Dry run: {args.dry_run}")
    log(f"[*] Candidate file path: {DEFAULT_PATH}")
    log(f"[*] Run By: {RUN_ACTOR}")
    log(f"[*] Source: {RUN_SOURCE}")

    if not DEFAULT_PATH.exists():
        log(f"[!] File not found: {DEFAULT_PATH}")
        sys.exit(1)

    if not USERNAME or not PASSWORD:
        log("[!] ERROR: Missing username/passwordAD environment variables.")
        sys.exit(1)

    lock_file = DEFAULT_PATH.parent / f"~${DEFAULT_PATH.name}"
    if lock_file.exists():
        owner = get_lock_owner(DEFAULT_PATH)
        log(f"[!] ABORTED: File is currently open by {owner}.")
        sys.exit(1)

    start_mtime = os.path.getmtime(DEFAULT_PATH)

    try:
        wb = load_workbook(DEFAULT_PATH, data_only=False)
        ws = wb.active
        log(f"[*] Workbook loaded: {DEFAULT_PATH}")
        log(f"[*] Active sheet: {ws.title}")
        log(f"[*] Max rows: {ws.max_row}, Max cols: {ws.max_column}")
    except Exception as exc:
        log(f"[!] Error loading workbook: {exc}")
        sys.exit(1)

    header_map = build_header_map(ws)
    log(f"[*] Header map: {header_map}")

    required = ("target_vlan", "switch_ip", "port", "mac", "ip", "last_checked", "notes")
    missing = [key for key in required if key not in header_map]
    if missing:
        log(f"[!] ERROR: Missing required headers: {missing}")
        sys.exit(1)

    # Group rows by switch to avoid reconnecting for every line
    rows_by_switch = defaultdict(list)

    for row_idx in range(4, ws.max_row + 1):
        switch_ip = clean_text(ws.cell(row=row_idx, column=header_map["switch_ip"]).value)
        port = clean_text(ws.cell(row=row_idx, column=header_map["port"]).value)
        target_vlan = clean_text(ws.cell(row=row_idx, column=header_map["target_vlan"]).value)

        if not switch_ip or not port or not target_vlan:
            continue

        rows_by_switch[switch_ip].append(
            {
                "row_idx": row_idx,
                "port": port,
                "target_vlan": target_vlan,
            }
        )

    if not rows_by_switch:
        log("[*] No usable rows found.")
        sys.exit(0)

    rows_checked = 0
    rows_already_correct = 0
    rows_changed = 0
    rows_failed = 0
    rows_declined = 0
    candidate_changes = 0
    workbook_touched = False

    config_summary = []

    if args.dry_run:
        log("!!!!!!!!!!!!!!!!!!!! DRY RUN ACTIVE !!!!!!!!!!!!!!!!!!!!")

    for switch_ip, row_entries in rows_by_switch.items():
        log(f"[*] Connecting to switch {switch_ip} for {len(row_entries)} row(s)")

        try:
            with connect_to_switch(switch_ip) as net_connect:
                prepare_switch_session(net_connect)

                live_vlan_map_before = get_live_vlan_map(net_connect)
                pending_verification = []

                for entry in row_entries:
                    row_idx = entry["row_idx"]
                    port = entry["port"]
                    target_vlan = entry["target_vlan"]
                    checked_at = now_str()

                    current_vlan = live_vlan_map_before.get(port, "Unknown")
                    live_details = get_port_live_details(net_connect, port)

                    rows_checked += 1

                    if current_vlan == target_vlan:
                        rows_already_correct += 1
                        log(
                            f"[OK] Row {row_idx} | {switch_ip} | Port {port} "
                            f"already on VLAN {target_vlan}"
                        )

                        if not args.dry_run:
                            write_readonly_columns(
                                ws,
                                row_idx,
                                header_map,
                                mac=live_details["mac"],
                                ip=live_details["ip"],
                                last_checked=checked_at,
                                notes="",
                            )
                            workbook_touched = True
                        continue

                    candidate_changes += 1

                    if args.dry_run:
                        log(
                            f"[DRY-RUN] Row {row_idx} | {switch_ip} | Port {port} | "
                            f"Live VLAN {current_vlan} -> Target VLAN {target_vlan}"
                        )
                        continue

                    try:
                        should_apply = confirm_change(
                            safe_mode=args.safe,
                            switch_ip=switch_ip,
                            port=port,
                            current_vlan=current_vlan,
                            target_vlan=target_vlan,
                            row_idx=row_idx,
                        )
                    except RuntimeError as exc:
                        log(f"[!] {exc}")
                        sys.exit(1)

                    if not should_apply:
                        rows_declined += 1
                        log(f"[SKIPPED] Row {row_idx} declined by user.")
                        write_readonly_columns(
                            ws,
                            row_idx,
                            header_map,
                            mac=live_details["mac"],
                            ip=live_details["ip"],
                            last_checked=checked_at,
                            notes=f"Current VLAN: {current_vlan}",
                        )
                        workbook_touched = True
                        continue

                    changed_by = RUN_ACTOR
                    changed_at = now_str()
                    source = RUN_SOURCE

                    log(
                        f"[*] Applying Row {row_idx} | "
                        f"Switch {switch_ip} | Port {port} | "
                        f"{current_vlan} -> {target_vlan} | "
                        f"By: {changed_by} | At: {changed_at} | Source: {source}"
                    )

                    try:
                        apply_vlan_change(net_connect, port, target_vlan)
                        pending_verification.append(
                            {
                                "row_idx": row_idx,
                                "port": port,
                                "target_vlan": target_vlan,
                                "old_vlan": current_vlan,
                                "changed_by": changed_by,
                                "changed_at": changed_at,
                                "source": source,
                            }
                        )
                    except Exception as exc:
                        rows_failed += 1
                        log(
                            f"[FAILED] Row {row_idx} | {switch_ip} | Port {port} | "
                            f"Error during config: {str(exc)[:200]}"
                        )
                        write_readonly_columns(
                            ws,
                            row_idx,
                            header_map,
                            mac=live_details["mac"],
                            ip=live_details["ip"],
                            last_checked=checked_at,
                            notes=f"Current VLAN: {current_vlan}",
                        )
                        workbook_touched = True

                # verify any attempted changes
                if pending_verification:
                    live_vlan_map_after = get_live_vlan_map(net_connect)

                    for item in pending_verification:
                        row_idx = item["row_idx"]
                        port = item["port"]
                        target_vlan = item["target_vlan"]
                        old_vlan = item["old_vlan"]
                        verified_vlan = live_vlan_map_after.get(port, "Unknown")
                        checked_at = now_str()
                        live_details = get_port_live_details(net_connect, port)

                        if verified_vlan == target_vlan:
                            rows_changed += 1
                            log(
                                f"[DONE] Row {row_idx} | {switch_ip} | Port {port} "
                                f"-> VLAN {target_vlan} | "
                                f"By: {item['changed_by']} | At: {item['changed_at']} | "
                                f"Source: {item['source']}"
                            )
                            write_readonly_columns(
                                ws,
                                row_idx,
                                header_map,
                                mac=live_details["mac"],
                                ip=live_details["ip"],
                                last_checked=checked_at,
                                notes="",
                            )
                            workbook_touched = True

                            config_summary.append(
                                {
                                    "ip": switch_ip,
                                    "port": port,
                                    "target_vlan": target_vlan,
                                    "old_vlan": old_vlan,
                                    "changed_by": item["changed_by"],
                                    "changed_at": item["changed_at"],
                                    "source": item["source"],
                                }
                            )
                        else:
                            rows_failed += 1
                            log(
                                f"[FAILED] Row {row_idx} | {switch_ip} | Port {port} | "
                                f"Target VLAN {target_vlan} not applied. Live VLAN is still {verified_vlan}"
                            )
                            write_readonly_columns(
                                ws,
                                row_idx,
                                header_map,
                                mac=live_details["mac"],
                                ip=live_details["ip"],
                                last_checked=checked_at,
                                notes=f"Current VLAN: {verified_vlan}",
                            )
                            workbook_touched = True

        except Exception as exc:
            rows_failed += len(row_entries)
            log(f"[!] Switch-level failure on {switch_ip}: {str(exc)[:250]}")

            if not args.dry_run:
                checked_at = now_str()
                for entry in row_entries:
                    write_readonly_columns(
                        ws,
                        entry["row_idx"],
                        header_map,
                        mac="Error",
                        ip="Error",
                        last_checked=checked_at,
                        notes="Switch connection failed",
                    )
                workbook_touched = True

    log(
        f"[*] Summary | Checked: {rows_checked} | Candidates: {candidate_changes} | "
        f"Changed: {rows_changed} | Already Correct: {rows_already_correct} | "
        f"Declined: {rows_declined} | Failed: {rows_failed}"
    )

    if args.dry_run:
        log("[*] Dry run finished successfully.")
        sys.exit(0)

    if workbook_touched:
        if os.path.getmtime(DEFAULT_PATH) != start_mtime:
            owner = get_lock_owner(DEFAULT_PATH)
            send_teams_notification(
                "CRITICAL",
                f"Conflict detected. {owner} modified the file while the script was running.",
            )
            log("[!] File changed while script was running. Save aborted.")
            sys.exit(1)

        try:
            wb.save(DEFAULT_PATH)
            log("[+] Success: Spreadsheet updated.")
        except PermissionError:
            owner = get_lock_owner(DEFAULT_PATH)
            log(f"[!] SAVE FAILED: {owner} has the file open.")
            sys.exit(1)
        except Exception as exc:
            log(f"[!] SAVE FAILED: {exc}")
            sys.exit(1)

    if rows_failed > 0:
        send_teams_notification(
            "WARNING",
            f"Completed with errors. Changed {rows_changed} port(s), but {rows_failed} row(s) failed.",
            details=config_summary,
        )
        sys.exit(1)

    if rows_changed > 0:
        send_teams_notification(
            "SUCCESS",
            f"Successfully updated {rows_changed} port(s).",
            details=config_summary,
        )
        sys.exit(0)

    log("[*] No VLAN changes needed.")
    sys.exit(0)


if __name__ == "__main__":
    main()
